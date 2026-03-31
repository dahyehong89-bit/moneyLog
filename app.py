import os
import re
import base64
import socket

from io import BytesIO
from datetime import date, datetime
from living import render_living_tab
from datetime import datetime, date
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
import calendar

FILE = "money.csv"
CHECKLIST_FILE = "checklist.csv"
COLUMNS = ["date", "amount", "category", "method", "memo"]

CATEGORY_OPTIONS = ["쇼핑", "외식", "배달", "커피", "편의점", "미용", "고정비", "사건비"]
METHOD_OPTIONS = ["현대카드", "신한카드", "사건비통장"]
DEFAULT_METHOD = "현대카드"
MONTHLY_BUDGET = 600000
BUDGET_METHOD = "현대카드"
PAGE_SIZE = 10

AUTO_CATEGORY = {
    "스타벅스": "커피",
    "커피": "커피",
    "카페": "커피",
    "배민": "배달",
    "요기요": "배달",
    "쿠팡이츠": "배달",
    "쿠팡": "쇼핑",
    "다이소": "쇼핑",
    "올리브영": "쇼핑",
    "편의점": "쇼핑",
    "마트": "쇼핑",
    "우유": "쇼핑",  
    "편의점": "편의점",
    "젤네일": "미용",
    "손젤": "미용",
    "발젤": "미용",
    "네일": "미용",
    "미용실": "미용",
}

AUTO_SHINHAN = [
    "주유",
    "통신비",
    "쿠팡와우",
    "이모티콘",
    "톨게이트",
]

INCIDENT_INCOME_KEYWORDS = ["환급", "입금", "수입", "보험금"]

CHECKLIST_ITEMS = [
    "공용 생활비 : 100만",
    "사건비 통장 : 20만",
    "수원 지역화폐 충전 : 10만",
    "모임비 : 19만",
    "정기 주차비 : 8만",
    "적금 : 30만",
    "청약 : 2만",
    "보험료1 : 23,226원",
    "보험료2 : 60,712원",
]

INCIDENT_CATEGORY_KEYWORDS = {
    "미용": ["네일", "젤", "손젤", "발젤", "젤네일", "미용실", "헤어", "컷", "염색"],
    "병원비": ["이비인후과", "내과", "소아과", "정형외과", "외과", "유방외과", "치과", "피부과", "안과", "산부인과", "병원", "의원", "진료", "외래", "감기"],
    "약값": ["약국", "처방", "약값"],
    "검진": ["검진", "건강검진", "초음파", "엑스레이", "x-ray", "mri", "ct"],
    "선물": ["선물", "생선", "생일선물", "축의", "축하", "꽃", "케이크"],
    "경조사": ["조의금", "부의금", "축의금", "결혼식", "장례식", "부고", "상가", "근조", "화환"]
}

INCIDENT_EXPENSE_KEYWORDS = {
    "미용": ["네일", "젤", "손젤", "발젤", "젤네일", "미용실", "헤어", "컷", "염색"],
    "병원비": ["이비인후과", "내과", "소아과", "정형외과", "외과", "유방외과", "치과", "피부과", "안과", "산부인과", "병원", "의원", "진료", "외래", "감기"],
    "약값": ["약국", "처방", "약값"],
    "검진": ["검진", "건강검진", "초음파", "엑스레이", "x-ray", "mri", "ct"],
    "선물": ["선물", "생일선물", "생선", "꽃", "케이크"],
    "경조사": ["조의금", "부의금", "축의금", "결혼식", "장례식", "부고", "상가", "근조", "화환"]    
}

# -----------------------
# 테마 설정
# -----------------------
THEMES = {
    "navy": {
        "app_bg_1": "#EEF2F8",
        "app_bg_2": "#E1E7F2",

        "container_bg": "rgba(255,255,255,0.65)",

        "input_border": "rgba(90,110,160,0.30)",
        "input_bg": "rgba(255,255,255,0.92)",

        "button_bg_1": "#A9B8E3",
        "button_bg_2": "#8FA4D6",
        "button_border": "rgba(90,110,160,0.35)",
        "button_text": "#2F416D",
        "button_shadow": "rgba(90,110,160,0.16)",

        "line": "rgba(90,110,160,0.18)",

        "metric_border": "rgba(90,110,160,0.20)",
        "expander_border": "rgba(90,110,160,0.18)",
        "form_border": "rgba(90,110,160,0.15)",

        "table_head_bg": "rgba(143,164,214,0.55)",
        "table_head_border": "rgba(90,110,160,0.25)",
        "table_head_text": "#3E568A",

        "row_hover": "rgba(143,164,214,0.20)",

        "amount_text": "#2F447A",

        "cat_bg": "rgba(143,164,214,0.30)",
        "cat_text": "#3E568A",

        "filter_bg": "rgba(255,255,255,0.90)",
        "filter_border": "rgba(90,110,160,0.30)",
        "filter_hover": "rgba(143,164,214,0.25)",

        "filter_active_1": "#A9B8E3",
        "filter_active_2": "#8FA4D6",
        "filter_active_border": "rgba(90,110,160,0.45)",
        "filter_shadow": "rgba(90,110,160,0.18)",

        "date_text": "#2F447A",
    },
}

@st.cache_resource
def get_gspread_client():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    credentials = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=scopes,
    )
    return gspread.authorize(credentials)


def get_worksheet(sheet_name: str):
    client = get_gspread_client()
    spreadsheet = client.open(st.secrets["sheets"]["spreadsheet_name"])
    return spreadsheet.worksheet(sheet_name)

@st.cache_data(ttl=60)
def load_df() -> pd.DataFrame:
    try:
        ws = get_worksheet("money")
        values = ws.get_all_records()

        if not values:
            return pd.DataFrame(columns=COLUMNS)

        df = pd.DataFrame(values).fillna("")

        rename_map = {
            "날짜": "date",
            "금액": "amount",
            "카테고리": "category",
            "결제수단": "method",
            "메모": "memo",
            "구분": "type",
        }

        df = df.rename(columns=rename_map)

        if "번호" in df.columns:
            df = df.drop(columns=["번호"])

        for c in ["date", "amount", "category", "method", "memo"]:
            if c not in df.columns:
                df[c] = ""

        if "type" not in df.columns:
            df["type"] = ""

        raw_amount = df["amount"].astype(str).str.strip()

        df["amount"] = (
            raw_amount
            .str.replace(",", "", regex=False)
            .str.replace("원", "", regex=False)
        )
        df["amount"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0).astype(int)

        df["type"] = df["type"].astype(str).str.strip()
        df["method"] = df["method"].astype(str).str.strip()
        df["memo"] = df["memo"].astype(str).str.strip()

        def restore_amount(row):
            amt = abs(int(row["amount"]))
            typ = row["type"]
            method = row["method"]
            memo = row["memo"]
            raw = str(row.get("_raw_amount", "")).strip()

            # 사건비통장은 메모 키워드 기준으로 우선 복원
            if method == "사건비통장":
                if any(k in memo for k in INCIDENT_INCOME_KEYWORDS):
                    return amt
                return -amt

            # 그 외는 구분 기준 복원
            if typ == "환급":
                return amt
            elif typ == "지출":
                return -amt

            # 예전 데이터 대응: 구분이 없으면 원래 부호 최대한 유지
            if raw.startswith("-"):
                return -amt
            elif raw.startswith("+"):
                return amt
            else:
                return -amt if method != "사건비통장" else amt

        df["_raw_amount"] = raw_amount
        df["amount"] = df.apply(restore_amount, axis=1)
        df = df.drop(columns=["_raw_amount"])

        df["date"] = df["date"].astype(str)
        df["category"] = df["category"].astype(str)
        df["method"] = df["method"].astype(str)
        df["memo"] = df["memo"].astype(str)

        return df[COLUMNS].copy()

    except Exception:
        return pd.DataFrame(columns=COLUMNS)

def save_df(df: pd.DataFrame) -> None:
    ws = get_worksheet("money")

    save_data = df[COLUMNS].copy().fillna("")

    # 날짜 정렬용
    save_data["date_dt"] = pd.to_datetime(save_data["date"], errors="coerce")

    save_data = save_data.sort_values(
        by=["date_dt", "date"],
        ascending=[False, False]
    ).drop(columns=["date_dt"])

    # 지출 / 환급 구분
    save_data["구분"] = save_data["amount"].apply(
        lambda x: "환급" if int(x) > 0 else "지출"
    )

    # 금액 표시
    save_data["금액"] = save_data["amount"].apply(
        lambda x: f"{abs(int(x)):,}원"
    )

    # 한글 컬럼 생성
    save_data["날짜"] = save_data["date"]
    save_data["카테고리"] = save_data["category"]
    save_data["결제수단"] = save_data["method"]
    save_data["메모"] = save_data["memo"]

    # 번호 생성
    save_data = save_data.reset_index(drop=True)
    save_data.insert(0, "번호", range(1, len(save_data) + 1))

    # 최종 컬럼 순서
    save_data = save_data[
        ["번호", "날짜", "구분", "카테고리", "메모", "금액", "결제수단"]
    ]

    rows = [save_data.columns.tolist()] + save_data.values.tolist()

    ws.clear()
    ws.update(rows)

    load_df.clear()

def get_month_sheet(gc, month_key):
    sh = gc.open("moneyLog")

    try:
        worksheet = sh.worksheet(month_key)
    except:
        worksheet = sh.add_worksheet(title=month_key, rows="1000", cols="20")

    return worksheet

def get_worksheet(sheet_name: str):
    client = get_gspread_client()
    spreadsheet = client.open(st.secrets["sheets"]["spreadsheet_name"])
    return spreadsheet.worksheet(sheet_name)

@st.cache_data(ttl=60)
def load_checklist_df() -> pd.DataFrame:
    try:
        ws = get_worksheet("checklist")
        values = ws.get_all_records()

        if not values:
            return pd.DataFrame(columns=["month", "item", "checked"])

        df = pd.DataFrame(values).fillna("")

        for c in ["month", "item", "checked"]:
            if c not in df.columns:
                df[c] = ""

        df = df[["month", "item", "checked"]].copy()
        df["month"] = df["month"].astype(str)
        df["item"] = df["item"].astype(str)
        df["checked"] = df["checked"].astype(str).map(
            lambda x: str(x).lower() in ["true", "1", "yes"]
        )

        return df

    except Exception:
        return pd.DataFrame(columns=["month", "item", "checked"])

@st.cache_data(ttl=60)
def load_no_spend_df() -> pd.DataFrame:
    try:
        ws = get_worksheet("no_spend_days")
        values = ws.get_all_records()

        if not values:
            return pd.DataFrame(columns=["date", "checked"])

        df = pd.DataFrame(values).fillna("")

        for c in ["date", "checked"]:
            if c not in df.columns:
                df[c] = ""

        df = df[["date", "checked"]].copy()
        df["date"] = df["date"].astype(str)
        df["checked"] = df["checked"].astype(str).map(
            lambda x: str(x).lower() in ["true", "1", "yes"]
        )

        return df

    except Exception:
        return pd.DataFrame(columns=["date", "checked"])


def save_no_spend_df(df: pd.DataFrame) -> None:
    ws = get_worksheet("no_spend_days")

    save_data = df[["date", "checked"]].copy().fillna("")
    rows = [["date", "checked"]] + save_data.values.tolist()

    ws.clear()
    ws.update(rows)
    load_no_spend_df.clear()

def update_no_spend_day(date_str: str, checked_value: bool) -> None:
    no_spend_df = load_no_spend_df()
    mask = no_spend_df["date"] == date_str

    if mask.any():
        no_spend_df.loc[mask, "checked"] = checked_value
    else:
        add_row = pd.DataFrame([{
            "date": date_str,
            "checked": checked_value,
        }])
        no_spend_df = pd.concat([no_spend_df, add_row], ignore_index=True)

    save_no_spend_df(no_spend_df)

def get_auto_no_spend_days(df: pd.DataFrame, month_key: str) -> set:
    if df.empty:
        return set()

    temp = df.copy()
    temp["date_dt"] = pd.to_datetime(temp["date"], errors="coerce")
    temp = temp.dropna(subset=["date_dt"]).copy()
    temp["month"] = temp["date_dt"].dt.strftime("%Y-%m")

    month_df = temp[temp["month"] == month_key].copy()
    if month_df.empty:
        return set()

    # 해당 월의 1일 ~ 오늘(현재월이면) / 월말(과거월이면) 범위
    selected_year, selected_month = map(int, month_key.split("-"))
    start_date = pd.Timestamp(year=selected_year, month=selected_month, day=1)

    KST = ZoneInfo("Asia/Seoul")
    today = pd.Timestamp(datetime.now(KST).date())

    if month_key == today.strftime("%Y-%m"):
        end_date = today
    else:
        end_date = (start_date + pd.offsets.MonthEnd(1))

    all_days = pd.date_range(start=start_date, end=end_date, freq="D")
    all_day_str = {d.strftime("%Y-%m-%d") for d in all_days}

    # 실지출 있는 날짜
    spend_days = set(
        month_df[month_df["amount"] < 0]["date"].astype(str).tolist()
    )

    # 실지출 없으면 자동 무지출
    auto_no_spend_days = all_day_str - spend_days
    return auto_no_spend_days

def get_final_no_spend_days(df: pd.DataFrame, month_key: str) -> set:
    auto_days = get_auto_no_spend_days(df, month_key)

    manual_df = load_no_spend_df()
    manual_df = manual_df[manual_df["checked"] == True].copy()

    manual_days = set()
    for d in manual_df["date"].astype(str):
        if d.startswith(month_key):
            manual_days.add(d)

    return auto_days | manual_days

def get_calendar_day_summary(df: pd.DataFrame, month_key: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["date", "spent", "refund", "has_fuel"])

    temp = df.copy()
    temp["date_dt"] = pd.to_datetime(temp["date"], errors="coerce")
    temp = temp.dropna(subset=["date_dt"]).copy()
    temp["month"] = temp["date_dt"].dt.strftime("%Y-%m")

    month_df = temp[temp["month"] == month_key].copy()
    if month_df.empty:
        return pd.DataFrame(columns=["date", "spent", "refund", "has_fuel"])

    summary = (
        month_df.groupby("date")
        .agg(
            spent=("amount", lambda s: int(s[s < 0].abs().sum())),
            refund=("amount", lambda s: int(s[s > 0].sum())),
            has_fuel=("memo", lambda s: bool(s.astype(str).str.contains("주유", na=False).any())),
        )
        .reset_index()
    )

    return summary


def render_month_calendar(df: pd.DataFrame, month_key: str, theme: dict):
    year, month = map(int, month_key.split("-"))

    day_summary_df = get_calendar_day_summary(df, month_key)
    no_spend_days = get_final_no_spend_days(df, month_key)

    summary_map = {}
    if not day_summary_df.empty:
        for _, row in day_summary_df.iterrows():
            summary_map[str(row["date"])] = {
                "spent": int(row["spent"]),
                "refund": int(row["refund"]),
                "has_fuel": bool(row["has_fuel"]),
            }

    cal = calendar.Calendar(firstweekday=0)  # 월요일 시작
    month_days = cal.monthdayscalendar(year, month)

    weekday_labels = ["월", "화", "수", "목", "금", "토", "일"]

    KST = ZoneInfo("Asia/Seoul")
    today_str = datetime.now(KST).strftime("%Y-%m-%d")
    current_month_str = datetime.now(KST).strftime("%Y-%m")

    html = f"""
    <style>
    .money-calendar-wrap {{
        display: grid;
        grid-template-columns: repeat(7, 1fr);
        gap: 8px;
        margin-top: 12px;
    }}

    .money-calendar-head {{
        background: {theme['table_head_bg']};
        border: 1px solid {theme['table_head_border']};
        color: {theme['table_head_text']};
        border-radius: 12px;
        padding: 8px 6px;
        text-align: center;
        font-weight: 800;
        font-size: 13px;
    }}

    .money-calendar-cell {{
        min-height: 112px;
        background: rgba(255,255,255,0.72);
        border: 1px solid {theme['metric_border']};
        border-radius: 16px;
        padding: 8px;
        box-sizing: border-box;
        display: flex;
        flex-direction: column;
        gap: 4px;
    }}

    .money-calendar-cell.empty {{
        background: rgba(255,255,255,0.28);
        border: 1px dashed {theme['metric_border']};
    }}

    .money-calendar-cell.today {{
        border: 2px solid {theme['filter_active_border']};
        box-shadow: 0 4px 14px {theme['filter_shadow']};
    }}

    .money-calendar-cell.no-spend {{
        background: rgba(238,249,241,0.95);
    }}

    .money-calendar-day {{
        font-weight: 800;
        font-size: 18px;
        color: {theme['button_text']};
        margin-bottom: 2px;
    }}

    .money-calendar-line {{
        font-size: 14px;
        line-height: 1.35;
        color: {theme['button_text']};
        word-break: keep-all;
    }}

    .money-calendar-line.spent {{
        color: #B4546A;
        font-weight: 700;
    }}

    .money-calendar-line.refund {{
        color: #2F7A4A;
        font-weight: 700;
    }}

    .money-calendar-line.nospend {{
        color: #2F7A4A;
        font-weight: 800;
    }}

    .money-calendar-line.fuel {{
        color: #6E56A0;
        font-weight: 700;
    }}

    @media (max-width: 768px) {{
        .money-calendar-wrap {{
            gap: 5px;
        }}

        .money-calendar-head {{
            font-size: 11px;
            padding: 6px 4px;
            border-radius: 10px;
        }}

        .money-calendar-cell {{
            min-height: 84px;
            border-radius: 12px;
            padding: 6px;
            gap: 2px;
        }}

        .money-calendar-day {{
            font-size: 15px;
        }}

        .money-calendar-line {{
            font-size: 12px;
            line-height: 1.2;
        }}
    }}
    </style>
    """

    html += '<div class="money-calendar-wrap">'

    for wd in weekday_labels:
        html += f'<div class="money-calendar-head">{wd}</div>'

    for week in month_days:
        for day in week:
            if day == 0:
                html += '<div class="money-calendar-cell empty"></div>'
                continue

            date_str = f"{year:04d}-{month:02d}-{day:02d}"
            info = summary_map.get(date_str, {"spent": 0, "refund": 0, "has_fuel": False})

            classes = ["money-calendar-cell"]
            if date_str == today_str and month_key == current_month_str:
                classes.append("today")
            if date_str in no_spend_days:
                classes.append("no-spend")

            lines = [f'<div class="money-calendar-day">{day}</div>']

            if info["spent"] > 0:
                lines.append(f'<div class="money-calendar-line spent">💸 {info["spent"]:,}원</div>')

            if info["refund"] > 0:
                lines.append(f'<div class="money-calendar-line refund">💰 {info["refund"]:,}원</div>')

            if info["has_fuel"]:
                lines.append(f'<div class="money-calendar-line fuel">⛽ 주유</div>')

            if date_str in no_spend_days:
                lines.append(f'<div class="money-calendar-line nospend">🪙 무지출</div>')

            html += f'<div class="{" ".join(classes)}">{"".join(lines)}</div>'

    html += "</div>"

    st.markdown(html, unsafe_allow_html=True)

def save_checklist_df(df: pd.DataFrame) -> None:
    ws = get_worksheet("checklist")

    save_data = df[["month", "item", "checked"]].copy().fillna("")
    rows = [["month", "item", "checked"]] + save_data.values.tolist()

    ws.clear()
    ws.update(rows)
    load_checklist_df.clear()

def get_month_checklist(month_key: str) -> pd.DataFrame:
    checklist_df = load_checklist_df()
    month_df = checklist_df[checklist_df["month"] == month_key].copy()

    if month_df.empty:
        month_df = pd.DataFrame({
            "month": [month_key] * len(CHECKLIST_ITEMS),
            "item": CHECKLIST_ITEMS,
            "checked": [False] * len(CHECKLIST_ITEMS),
        })
        checklist_df = pd.concat([checklist_df, month_df], ignore_index=True)
        save_checklist_df(checklist_df)
    else:
        existing_items = set(month_df["item"].tolist())
        missing_items = [x for x in CHECKLIST_ITEMS if x not in existing_items]

        if missing_items:
            add_df = pd.DataFrame({
                "month": [month_key] * len(missing_items),
                "item": missing_items,
                "checked": [False] * len(missing_items),
            })
            checklist_df = pd.concat([checklist_df, add_df], ignore_index=True)
            save_checklist_df(checklist_df)
            month_df = checklist_df[checklist_df["month"] == month_key].copy()

    return month_df

def parse_checklist_amount(item_text: str) -> int:
    text = str(item_text or "").replace(",", "").strip()

    m = re.search(r"(\d+)\s*만", text)
    if m:
        return int(m.group(1)) * 10000

    m = re.search(r"(\d+)\s*원", text)
    if m:
        return int(m.group(1))

    return 0

def update_checklist_item(month_key: str, item_name: str, checked_value: bool) -> None:
    checklist_df = load_checklist_df()
    mask = (checklist_df["month"] == month_key) & (checklist_df["item"] == item_name)

    if mask.any():
        checklist_df.loc[mask, "checked"] = checked_value
    else:
        add_row = pd.DataFrame([{
            "month": month_key,
            "item": item_name,
            "checked": checked_value,
        }])
        checklist_df = pd.concat([checklist_df, add_row], ignore_index=True)

    save_checklist_df(checklist_df)


def auto_category_from_text(text: str, fallback: str = "쇼핑") -> str:
    t = (text or "").strip()

    for k, v in AUTO_CATEGORY.items():
        if k and k in t:
            return v

    return fallback


def auto_card_and_category(text: str, default_category: str, default_method: str):
    t = (text or "").strip()

    # 신한카드 고정비 우선
    for k in AUTO_SHINHAN:
        if k in t:
            return "고정비", "신한카드"

    # 사건비통장 자동 분류
    if is_incident_expense_text(t):
        return "사건비", "사건비통장"

    category = auto_category_from_text(t, default_category)
    return category, default_method

def is_incident_expense_text(text: str) -> bool:
    t = (text or "").strip()

    incident_only_categories = ["병원비", "약값", "검진", "선물", "경조사"]

    for category in incident_only_categories:
        keywords = INCIDENT_EXPENSE_KEYWORDS.get(category, [])
        for keyword in keywords:
            if keyword in t:
                return True

    return False

def split_fuel_memo(memo: str):
    memo = (memo or "").strip()

    fuel_price = ""
    actual_amount = ""
    is_non_expense = "[비지출]" in memo

    fuel_match = re.search(r"리터당\s*([\d,]+)원", memo)
    actual_match = re.search(r"실제\s*([\d,]+)원", memo)

    if fuel_match:
        fuel_price = fuel_match.group(1).replace(",", "")

    if actual_match:
        actual_amount = actual_match.group(1).replace(",", "")

    clean_memo = memo
    clean_memo = re.sub(r"\s*/?\s*\[비지출\]", "", clean_memo).strip()
    clean_memo = re.sub(r"\s*/?\s*실제\s*[\d,]+원", "", clean_memo).strip()
    clean_memo = re.sub(r"\s*/?\s*리터당\s*[\d,]+원", "", clean_memo).strip()
    clean_memo = re.sub(r"\s*/?\s*[\d.]+L", "", clean_memo).strip()
    clean_memo = re.sub(r"\s*/\s*", " / ", clean_memo).strip(" /")

    return clean_memo, fuel_price, actual_amount, is_non_expense

def is_incident_income(method: str, memo: str) -> bool:
    if method != "사건비통장":
        return False

    memo_text = (memo or "").strip()
    return any(k in memo_text for k in INCIDENT_INCOME_KEYWORDS)


def build_fuel_memo(
    memo: str,
    fuel_price_clean: str,
    amount_clean: str,
    is_non_expense: bool = False
) -> str:
    final_memo = (memo or "").strip()

    if fuel_price_clean:
        fuel_price_int = int(fuel_price_clean)
        amount_int = int(amount_clean)

        liters_str = ""
        if fuel_price_int > 0 and amount_int > 0:
            liters = amount_int / fuel_price_int
            liters_str = f"{liters:.2f}L"

        parts = []

        if final_memo:
            parts.append(final_memo)
        else:
            parts.append("주유")

        if is_non_expense:
            parts.append("[비지출]")
            parts.append(f"실제 {amount_int:,}원")

        parts.append(f"리터당 {fuel_price_int:,}원")

        if liters_str:
            parts.append(liters_str)

        return " / ".join(parts)

    if is_non_expense:
        if final_memo:
            return f"{final_memo} / [비지출]"
        return "[비지출]"

    return final_memo

def classify_incident_memo(memo: str) -> str:
    memo_text = str(memo or "").strip().lower()
    memo_text_no_space = memo_text.replace(" ", "")

    all_pairs = []
    for category, keywords in INCIDENT_CATEGORY_KEYWORDS.items():
        for keyword in keywords:
            kw = str(keyword).strip().lower()
            all_pairs.append((kw, category))

    # 긴 키워드부터 먼저 검사
    all_pairs.sort(key=lambda x: len(x[0]), reverse=True)

    for keyword, category in all_pairs:
        keyword_no_space = keyword.replace(" ", "")

        if keyword in memo_text or keyword_no_space in memo_text_no_space:
            return category

    return "기타"

def extract_fuel_stats_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["date_dt", "date", "unit_price", "fuel_amount", "liters"])

    fuel_df = df.copy()
    fuel_df["memo"] = fuel_df["memo"].astype(str)
    fuel_df["date_dt"] = pd.to_datetime(fuel_df["date"], errors="coerce")

    # 리터당 가격
    fuel_df["unit_price"] = fuel_df["memo"].str.extract(r"리터당\s*([\d,]+)원")[0]
    fuel_df["unit_price"] = (
        fuel_df["unit_price"]
        .astype(str)
        .str.replace(",", "", regex=False)
    )
    fuel_df["unit_price"] = pd.to_numeric(fuel_df["unit_price"], errors="coerce")

    # 리터
    fuel_df["liters"] = fuel_df["memo"].str.extract(r"([\d.]+)L")[0]
    fuel_df["liters"] = pd.to_numeric(fuel_df["liters"], errors="coerce")

    # 실제 금액 (비지출 주유용)
    fuel_df["actual_amount"] = fuel_df["memo"].str.extract(r"실제\s*([\d,]+)원")[0]
    fuel_df["actual_amount"] = (
        fuel_df["actual_amount"]
        .astype(str)
        .str.replace(",", "", regex=False)
    )
    fuel_df["actual_amount"] = pd.to_numeric(fuel_df["actual_amount"], errors="coerce")

    # 실제 주유금액 계산
    # 비지출이면 memo의 실제 금액 사용, 아니면 amount 절대값 사용
    fuel_df["fuel_amount"] = fuel_df["actual_amount"]
    fuel_df.loc[fuel_df["fuel_amount"].isna(), "fuel_amount"] = fuel_df["amount"].abs()

    # liters 없으면 금액 / 단가로 계산
    missing_liters = (
        fuel_df["liters"].isna()
        & fuel_df["unit_price"].notna()
        & (fuel_df["unit_price"] > 0)
        & fuel_df["fuel_amount"].notna()
    )
    fuel_df.loc[missing_liters, "liters"] = (
        fuel_df.loc[missing_liters, "fuel_amount"] / fuel_df.loc[missing_liters, "unit_price"]
    )

    fuel_df = fuel_df.dropna(subset=["date_dt"]).sort_values("date_dt")

    return fuel_df[["date_dt", "date", "unit_price", "fuel_amount", "liters"]].copy()

def parse_quick_input(text: str, default_category: str, default_method: str) -> dict:
    text = (text or "").strip()
    if not text:
        raise ValueError("빈 입력이에요.")

    tokens = text.split()
    if not tokens:
        raise ValueError("입력값이 없어요.")

    d = str(datetime.now(ZoneInfo("Asia/Seoul")).date())
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", tokens[0]):
        d = tokens[0]
        tokens = tokens[1:]
        if not tokens:
            raise ValueError("날짜만 있고 금액이 없어요. 예: 2026-03-01 우유 4500")

    method = default_method or DEFAULT_METHOD
    filtered_tokens = []

    for t in tokens:
        if t.startswith("@") and len(t) > 1:
            method = t[1:]
        else:
            filtered_tokens.append(t)

    amount = None
    memo_tokens = []

    for t in filtered_tokens:
        t_clean = t.replace(",", "")
        if amount is None and re.fullmatch(r"[+-]?\d+", t_clean):
            amount = int(t_clean)
        else:
            memo_tokens.append(t)

    if amount is None:
        raise ValueError("금액이 없어요. 예: 4500 우유 / 우유 4500")

    memo = " ".join(memo_tokens).strip()
    category, method = auto_card_and_category(memo, default_category, method)

    # 사건비통장 + 환급/입금/수입/보험금 => 수입(+)
    if method == "사건비통장" and any(k in memo for k in INCIDENT_INCOME_KEYWORDS):
        final_amount = abs(amount)
    else:
        final_amount = -abs(amount)

    return {
        "date": d,
        "amount": final_amount,
        "category": category,
        "method": method or DEFAULT_METHOD,
        "memo": memo,
    }


def get_month_options(df: pd.DataFrame):
    KST = ZoneInfo("Asia/Seoul")
    current_month = datetime.now(KST).strftime("%Y-%m")

    if df.empty:
        return [current_month]

    dts = pd.to_datetime(df["date"], errors="coerce")
    months = sorted(
        {m.strftime("%Y-%m") for m in dts.dropna().dt.to_period("M").dt.to_timestamp()},
        reverse=True
    )

    if current_month not in months:
        months.insert(0, current_month)

    return months or [current_month]


def load_font():
    if not os.path.exists("leejieun.ttf"):
        return

    with open("leejieun.ttf", "rb") as f:
        font_bytes = f.read()

    encoded = base64.b64encode(font_bytes).decode()

    st.markdown(
        f"""
        <style>
        @font-face {{
            font-family: 'LeeJieun';
            src: url(data:font/ttf;base64,{encoded}) format('truetype');
            font-weight: normal;
            font-style: normal;
        }}
        </style>
        """,
        unsafe_allow_html=True
    )

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("10.255.255.255", 1))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return None


def budget_status(spent: int):
    ratio = spent / MONTHLY_BUDGET if MONTHLY_BUDGET > 0 else 0

    if ratio >= 1:
        return {
            "label": "초과",
            "bg": "#FFE3E8",
            "border": "#FF7A9A",
            "text": "#C33B5E",
        }
    elif ratio >= 0.8:
        return {
            "label": "거의 다 씀",
            "bg": "#FFF1E0",
            "border": "#FFB85C",
            "text": "#C97A00",
        }
    elif ratio >= 0.5:
        return {
            "label": "주의",
            "bg": "#FFF8D9",
            "border": "#E7C84C",
            "text": "#9A7A00",
        }
    else:
        return {
            "label": "여유",
            "bg": "#E9F9EF",
            "border": "#72D69B",
            "text": "#257A45",
        }

def get_budget_review(spent: int, budget: int = MONTHLY_BUDGET):
    diff = budget - spent

    if diff < 0:
        return {
            "result_text": f"{abs(diff):,}원 초과",
            "comment": "조금 오버했어요 🥲 다음 달엔 더 잘해보자!",
            "bg": "#FFF1F3",
            "border": "#F3B7C3",
            "text": "#B4546A",
        }

    elif diff == 0:
        return {
            "result_text": "예산 딱 맞춤",
            "comment": "와… 이건 진짜 계획형 인간 😮✨",
            "bg": "#F8F4FF",
            "border": "#D8C7F3",
            "text": "#6E56A0",
        }

    else:
        return {
            "result_text": f"{diff:,}원 남김",
            "comment": "잘 참았다 👏 예산 세이브 성공!",
            "bg": "#EEF9F1",
            "border": "#BFE5C8",
            "text": "#2F7A4A",
        }

def get_monthly_budget_reviews(df: pd.DataFrame, months: int = 3) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["month", "spent", "diff", "result_text", "comment"])

    temp = df.copy()
    temp["date_dt"] = pd.to_datetime(temp["date"], errors="coerce")
    temp = temp.dropna(subset=["date_dt"]).copy()

    # 예산 기준은 현대카드만
    temp = temp[temp["method"] == BUDGET_METHOD].copy()

    if temp.empty:
        return pd.DataFrame(columns=["month", "spent", "diff", "result_text", "comment"])

    temp["month"] = temp["date_dt"].dt.strftime("%Y-%m")

    from datetime import datetime
    from zoneinfo import ZoneInfo

    current_month = datetime.now(ZoneInfo("Asia/Seoul")).strftime("%Y-%m")
    temp = temp[temp["month"] != current_month]

    month_sum = (
        temp.groupby("month")["amount"]
        .sum()
        .reset_index()
    )

    month_sum["spent"] = month_sum["amount"].abs().astype(int)
    month_sum = month_sum.sort_values("month", ascending=False).head(months).copy()

    results = []
    for _, row in month_sum.iterrows():
        spent = int(row["spent"])
        review = get_budget_review(spent, MONTHLY_BUDGET)

        results.append({
            "month": row["month"],
            "spent": spent,
            "diff": MONTHLY_BUDGET - spent,
            "result_text": review["result_text"],
            "comment": review["comment"],
            "bg": review["bg"],
            "border": review["border"],
            "text": review["text"],
        })

    return pd.DataFrame(results)

def render_budget_card(title: str, value: str, bg: str, border: str, text: str):
    st.markdown(
        f"""
        <div style="
            background:{bg};
            border:1px solid {border};
            border-radius:18px;
            padding:16px;
            min-height:105px;
            box-shadow:0 6px 18px rgba(0,0,0,0.04);
        ">
            <div style="font-size:14px; color:{text}; margin-bottom:8px;">{title}</div>
            <div style="font-size:28px; font-weight:800; color:{text};">{value}</div>
        </div>
        """,
        unsafe_allow_html=True
    )


# -----------------------------
# App
# -----------------------------
st.set_page_config(page_title="빠른 가계부", layout="wide")
df = load_df()
load_font()

theme_name = st.session_state.get("theme_select", "navy")
theme = THEMES[theme_name]

st.markdown(f"""
<style>
/* ===== 폰트 ===== */
html, body, [data-testid="stApp"], [data-testid="stAppViewContainer"],
[data-testid="stSidebar"], div, p, span, label, li {{
    font-family: "LeeJieun", "Segoe UI Emoji", "Apple Color Emoji", sans-serif;
}}

input,
textarea,
button,
label,
select,
input[type="text"],
input[type="number"],
input[type="date"] {{
    font-family: "LeeJieun", "Segoe UI Emoji", "Apple Color Emoji", sans-serif !important;
}}

div[data-testid="stTextInput"] input,
div[data-testid="stNumberInput"] input,
div[data-testid="stDateInput"] input,
div[data-testid="stTextArea"] textarea,
[data-baseweb="select"] *,
div[data-testid="stSelectbox"] * {{
    font-family: "LeeJieun", "Segoe UI Emoji", "Apple Color Emoji", sans-serif !important;
}}

/* ===== 배경 ===== */
html, body, [data-testid="stAppViewContainer"] {{
    background: linear-gradient(180deg, {theme['app_bg_1']} 0%, {theme['app_bg_2']} 100%) !important;
}}

[data-testid="stAppViewContainer"] > .main {{
    background: transparent !important;
}}

/* ===== 메인 컨테이너 ===== */
.block-container {{
    padding-top: 1rem !important;
    padding-bottom: 1.2rem !important;
    padding-left: 0.8rem !important;
    padding-right: 0.8rem !important;
    max-width: 1400px;
    background: {theme['container_bg']};
    border-radius: 24px;
}}

/* ===== 타이틀 ===== */
h1, h2, h3 {{
    font-weight: 800;
    letter-spacing: -1px;
}}

h3, h4 {{
    margin-top: 0.3rem !important;
    margin-bottom: 0.5rem !important;
}}

hr {{
    border: none;
    height: 1px;
    background: {theme['line']};
    margin: 1.1rem 0;
}}

/* ===== 입력창 ===== */
input, textarea {{
    border-radius: 14px !important;
    border: 1px solid {theme['input_border']} !important;
    background: {theme['input_bg']} !important;
}}

[data-baseweb="select"] > div {{
    border-radius: 14px !important;
    border: 1px solid {theme['input_border']} !important;
    background: {theme['input_bg']} !important;
    cursor: pointer !important;
}}

/* ===== 기본 버튼 ===== */
button:not([kind="primary"]) {{
    border-radius: 16px !important;
    border: 1px solid {theme['button_border']} !important;
    background: linear-gradient(180deg, {theme['button_bg_1']}, {theme['button_bg_2']}) !important;
    color: {theme['button_text']} !important;
    font-weight: 700 !important;
    box-shadow: 0 10px 18px {theme['button_shadow']} !important;
    transition: all 0.15s ease !important;
}}

button:hover {{
    transform: translateY(-1px);
}}

/* ===== 필터 버튼 상태 ===== */
button[kind="secondary"] {{
    background: {theme['filter_bg']} !important;
    border: 1px solid {theme['filter_border']} !important;
    color: {theme['button_text']} !important;
    box-shadow: none !important;
}}

button[kind="secondary"]:hover {{
    background: {theme['filter_hover']} !important;
    border: 1px solid {theme['filter_active_border']} !important;
    transform: translateY(-1px);
}}

button[kind="primary"] {{
    background: linear-gradient(180deg, {theme['filter_active_1']}, {theme['filter_active_2']}) !important;
    border: 1px solid {theme['filter_active_border']} !important;
    color: {theme['button_text']} !important;
    box-shadow: 0 8px 16px {theme['filter_shadow']} !important;
    font-weight: 800 !important;
}}

button[kind="primary"]:hover {{
    background: linear-gradient(180deg, {theme['filter_active_1']}, {theme['filter_active_2']}) !important;
    border: 1px solid {theme['filter_active_border']} !important;
}}

/* ===== metric / form ===== */
[data-testid="stMetric"] {{
    background: rgba(255,255,255,0.60);
    border-radius: 16px;
    padding: 10px;
    border: 1px solid {theme['metric_border']};
}}

[data-testid="stForm"] {{
    background: rgba(255,255,255,0.45);
    border: 1px solid {theme['form_border']};
    border-radius: 18px;
    padding: 14px;
}}

/* ===== 테이블 ===== */
.table-head {{
    font-weight: 800;
    background: {theme['table_head_bg']};
    border: 1px solid {theme['table_head_border']};
    border-radius: 12px;
    padding: 8px 10px;
    text-align: center;
    color: {theme['table_head_text']};
    margin-bottom: 4px;
    font-size: 14px;
}}

.row-box {{
    padding: 6px;
    border-radius: 10px;
    text-align: center;
    font-size: 14px;
    line-height: 1.35;
}}

.row-box:hover {{
    background: {theme['row_hover']};
}}

.amount-text {{
    font-weight: 800;
    color: {theme['amount_text']};
}}

/* ===== 태그 ===== */
.cat-tag {{
    background: {theme['cat_bg']};
    color: {theme['cat_text']};
    padding: 4px 10px;
    border-radius: 999px;
    font-weight: 700;
    font-size: 15px;
    display: inline-block;
}}

.method-hyundai {{
    background: rgba(200,200,200,0.35);
    color: #4A4A4A;
    padding: 4px 10px;
    border-radius: 999px;
    font-weight: 700;
    font-size: 15px;
    display: inline-block;
}}

.method-shinhan {{
    background: rgba(180,220,255,0.45);
    color: #2F6F8F;
    padding: 4px 10px;
    border-radius: 999px;
    font-weight: 700;
    font-size: 15px;
    display: inline-block;
}}

.method-incident {{
    background: rgba(255,225,120,0.45);
    color: #8A6A00;
    padding: 4px 10px;
    border-radius: 999px;
    font-weight: 700;
    font-size: 15px;
    display: inline-block;
}}

/* ===== 버튼 간격 ===== */
div[data-testid="stButton"] {{
    margin-top: 0 !important;
    margin-bottom: 0 !important;
}}

div[data-testid="stButton"] button {{
    padding: 2px 6px !important;
    min-height: 30px !important;
    border-radius: 16px;
}}

div[data-testid="stButton"] button:hover {{
    background: {theme['filter_hover']} !important;
}}

/* 👇 클릭된 상태 느낌 */
div[data-testid="stButton"] button[kind="secondary"]:focus {{
    background: {theme['filter_active_1']} !important;
    box-shadow: 0 0 0 2px {theme['filter_active_border']} !important;
}}

/* ===== 체크박스 ===== */
label:has(input[type="checkbox"]) {{
    font-weight: 700;
}}

/* ===== radio group ===== */
div[role="radiogroup"] {{
    gap: 8px !important;
    padding-left: 0px !important;
    margin-left: 0px !important;
}}

div[role="radiogroup"] * {{
    box-shadow: none !important;
}}

div[role="radiogroup"] label {{
    background: {theme['filter_bg']} !important;
    border: 1px solid {theme['filter_border']} !important;
    border-radius: 14px !important;
    padding: 8px 14px !important;
    min-height: 42px !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    cursor: pointer !important;
    transition: all 0.15s ease !important;
    margin-left: 0px !important;
}}

div[role="radiogroup"] label:hover {{
    background: {theme['filter_hover']} !important;
    transform: translateY(-1px);
}}

div[role="radiogroup"] label:has(input:checked) {{
    background: linear-gradient(180deg, {theme['filter_active_1']}, {theme['filter_active_2']}) !important;
    border: 1px solid {theme['filter_active_border']} !important;
    box-shadow: 0 8px 16px {theme['filter_shadow']} !important;
    color: {theme['button_text']} !important;
}}

div[role="radiogroup"] label > div:first-child {{
    display: none !important;
}}

div[role="radiogroup"] input[type="radio"] {{
    display: none !important;
}}

div[role="radiogroup"] input[type="radio"],
div[role="radiogroup"] input[type="radio"] + div,
div[role="radiogroup"] [data-testid="stMarkdownContainer"] + div {{
    accent-color: transparent !important;
}}

/* ===== 탭 ===== */
button[data-baseweb="tab"] {{
    background: {theme['filter_bg']} !important;
    border: 1px solid {theme['filter_border']} !important;
    border-radius: 14px !important;
    color: {theme['button_text']} !important;
    font-weight: 700 !important;
    padding: 10px 18px !important;
    height: auto !important;
    transition: all 0.15s ease !important;
    box-shadow: none !important;
    margin-right: 8px !important;
}}

button[data-baseweb="tab"]:hover {{
    background: {theme['filter_hover']} !important;
    transform: translateY(-1px);
}}

button[data-baseweb="tab"][aria-selected="true"] {{
    background: linear-gradient(180deg, {theme['filter_active_1']}, {theme['filter_active_2']}) !important;
    border: 1px solid {theme['filter_active_border']} !important;
    color: {theme['button_text']} !important;
    box-shadow: 0 8px 16px {theme['filter_shadow']} !important;
}}

div[data-baseweb="tab-border"] {{
    background: transparent !important;
}}

div[data-baseweb="tab-highlight"] {{
    background: transparent !important;
}}

/* ===== 진행바 ===== */
[data-testid="stProgress"] [data-baseweb="progress-bar"] > div {{
    background: rgba(255,255,255,0.55) !important;
    border: 1px solid {theme['metric_border']} !important;
    border-radius: 999px !important;
    overflow: hidden !important;
}}

[data-testid="stProgress"] [role="progressbar"] div[style*="width"],
[data-testid="stProgress"] [data-baseweb="progress-bar"] div div div {{
    background: linear-gradient(90deg, {theme['button_bg_1']}, {theme['button_bg_2']}) !important;
    border-radius: 999px !important;
}}

div[data-testid="stDataFrame"] * {{
    font-family: "LeeJieun", "Segoe UI Emoji", "Apple Color Emoji", sans-serif !important;
}}

div[data-testid="stDataFrameGlideDataEditor"] * {{
    font-family: "LeeJieun", "Segoe UI Emoji", "Apple Color Emoji", sans-serif !important;
}}

/* ===== 모바일 전용 ===== */
@media (max-width: 768px) {{
    .block-container {{
        padding-top: 0.55rem !important;
        padding-bottom: 0.8rem !important;
        padding-left: 0.45rem !important;
        padding-right: 0.45rem !important;
        border-radius: 16px !important;
    }}

    h1, h2, h3 {{
        letter-spacing: -0.5px !important;
    }}

    h3 {{
        font-size: 1.02rem !important;
        margin-top: 0.15rem !important;
        margin-bottom: 0.35rem !important;
    }}

    h4 {{
        font-size: 0.95rem !important;
        margin-top: 0.1rem !important;
        margin-bottom: 0.25rem !important;
    }}

    hr {{
        margin: 0.65rem 0 !important;
    }}

    label[data-testid="stWidgetLabel"] p {{
        font-size: 0.82rem !important;
        margin-bottom: 0.05rem !important;
    }}

    input, textarea {{
        font-size: 14px !important;
    }}

    [data-baseweb="select"] > div {{
        min-height: 2.5rem !important;
    }}

    div[data-testid="stTextInput"],
    div[data-testid="stDateInput"],
    div[data-testid="stSelectbox"] {{
        margin-bottom: 0.08rem !important;
    }}

    [data-testid="stForm"] {{
        padding: 10px !important;
        border-radius: 14px !important;
    }}

    [data-testid="stMetric"] {{
        padding: 8px !important;
        border-radius: 12px !important;
    }}

    .table-head {{
        font-size: 11px !important;
        padding: 6px 6px !important;
        border-radius: 10px !important;
    }}

    .row-box {{
        padding: 5px 3px !important;
        font-size: 11.5px !important;
        line-height: 1.2 !important;
        word-break: keep-all;
    }}

    .amount-text {{
        font-size: 11.5px !important;
    }}

    .cat-tag,
    .method-hyundai,
    .method-shinhan,
    .method-incident {{
        font-size: 11px !important;
        padding: 3px 7px !important;
    }}

    div[data-testid="stButton"] button {{
        min-height: 34px !important;
        font-size: 13px !important;
        padding: 4px 8px !important;
        border-radius: 12px !important;
    }}

    button[data-baseweb="tab"] {{
        font-size: 12px !important;
        padding: 8px 12px !important;
        margin-right: 6px !important;
        border-radius: 12px !important;
    }}

    div[role="radiogroup"] {{
        gap: 6px !important;
    }}

    div[role="radiogroup"] label {{
        min-height: 36px !important;
        padding: 6px 10px !important;
        font-size: 12px !important;
        border-radius: 12px !important;
    }}

    div[data-testid="stMarkdownContainer"]{{
        margin-bottom:0;
    }}
}}
</style>
""", unsafe_allow_html=True)

st.markdown("<div style='margin-top:20px;'></div>", unsafe_allow_html=True)

# -------------------
# 상단 제목 + 오늘 날짜
# -------------------
weekday = ["월", "화", "수", "목", "금", "토", "일"]
KST = ZoneInfo("Asia/Seoul")
today = datetime.now(KST)
today_str = f"{today.strftime('%Y.%m.%d')} ({weekday[today.weekday()]})"

title_left, title_right = st.columns([3, 1])

with title_left:
    st.title("💸 빠른 가계부")

with title_right:
    st.markdown(
        f"""
        <div style="text-align:right;margin-top:10px;">
            <div style="font-size:14px;color:{theme['date_text']};font-weight:700;">
                오늘
            </div>
            <div style="font-size:20px;font-weight:800;color:{theme['date_text']};">
                {today_str}
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

# 탭
tab1, tab2, tab3 = st.tabs(["🏠 개인 가계부", "📊 내역", "🏦 생활비 통장"])

# -------------------
# 수정 모달
# -------------------
@st.dialog("✏ 기록 수정")
def edit_dialog(rid: int):
    current_df = load_df()

    if rid >= len(current_df):
        st.error("수정할 데이터를 찾지 못했어요.")
        return

    row = current_df.iloc[rid]
    current_cat = str(row["category"])
    cat_index = CATEGORY_OPTIONS.index(current_cat) if current_cat in CATEGORY_OPTIONS else 0

    current_method = str(row["method"]).strip() if str(row["method"]).strip() else DEFAULT_METHOD
    method_index = METHOD_OPTIONS.index(current_method) if current_method in METHOD_OPTIONS else 0

    base_memo, base_fuel_price, base_actual_amount, base_is_non_expense = split_fuel_memo(str(row["memo"]))

    with st.form(f"edit_form_{rid}"):
        q1, q2 = st.columns(2)

        with q1:
            category = st.selectbox(
                "카테고리",
                CATEGORY_OPTIONS,
                index=cat_index,
                key=f"edit_cat_{rid}"
            )

            d = st.date_input(
                "날짜",
                value=pd.to_datetime(row["date"], errors="coerce"),
                key=f"edit_date_{rid}"
            )

            method = st.selectbox(
                "결제수단",
                METHOD_OPTIONS,
                index=method_index,
                key=f"edit_method_{rid}",
                disabled=False
            )

        with q2:
            memo = st.text_input(
                "메모",
                value=base_memo,
                key=f"edit_memo_{rid}"
            )

            default_amount = base_actual_amount if base_is_non_expense and base_actual_amount else f"{abs(int(row['amount']))}"
            amount = st.text_input(
                "금액",
                value=f"{int(default_amount):,}" if str(default_amount).isdigit() else "",
                key=f"edit_amount_{rid}"
            )

            fuel_price = st.text_input(
                "리터당 가격",
                value=base_fuel_price,
                placeholder="주유일 때만 입력", 
                key=f"edit_fuel_price_{rid}"
            )

        c_left, c_right = st.columns([1, 1])

        with c_right:
            non_expense = st.checkbox(
                "지출에 반영 안 함 (지출 제외)",
                value=base_is_non_expense,
                key=f"edit_non_expense_{rid}"
            )

            if non_expense:
                st.caption("비지출 기록은 결제수단을 저장하지 않아요.")
                method = ""

        col_cancel, col_save = st.columns(2)

        with col_cancel:
            canceled = st.form_submit_button("취소", use_container_width=True)

        with col_save:
            saved = st.form_submit_button("💾 저장", use_container_width=True)

    if saved:
        amount_clean = amount.replace(",", "").strip()
        fuel_price_clean = fuel_price.replace(",", "").strip()

        if not amount_clean or not re.fullmatch(r"\d+", amount_clean):
            st.error("금액은 숫자만 입력해줘.")
        elif fuel_price_clean and not re.fullmatch(r"\d+", fuel_price_clean):
            st.error("리터당 가격은 숫자만 입력해줘.")
        else:
            final_memo = build_fuel_memo(
                memo,
                fuel_price_clean,
                amount_clean,
                is_non_expense=non_expense
            )

            if non_expense:
                final_category = category
                final_method = ""
            else:
                final_category, final_method = auto_card_and_category(
                    final_memo,
                    category,
                    method or DEFAULT_METHOD
                )

            amount_value = int(amount_clean)

            if non_expense:
                final_amount = 0
            elif is_incident_income(final_method, final_memo):
                final_amount = amount_value
            else:
                final_amount = -amount_value

            current_df.iloc[rid] = [
                str(d),
                final_amount,
                final_category,
                final_method,
                final_memo
            ]
            save_df(current_df)
            st.success("수정 완료!")
            st.rerun()

    if canceled:
        st.rerun()

if "pending_quick_entry" not in st.session_state:
    st.session_state.pending_quick_entry = None


def add_quick(amount: int, category: str, memo: str = "", method: str = DEFAULT_METHOD):
    global df

    amount = abs(int(amount))
    final_category, final_method = auto_card_and_category(memo, category, method)

    if is_incident_income(final_method, memo):
        final_amount = amount
    else:
        final_amount = -amount

    row = {
        "date": str(datetime.now(ZoneInfo("Asia/Seoul")).date()),
        "amount": final_amount,
        "category": final_category,
        "method": final_method,
        "memo": memo,
    }

    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    save_df(df)
    st.success("✅ 저장 완료!")
    st.rerun()


def open_quick_edit(
    amount: int,
    category: str,
    memo: str = "",
    method: str = DEFAULT_METHOD,
    non_expense: bool = False
):
    st.session_state.pending_quick_entry = {
        "date": datetime.now(ZoneInfo("Asia/Seoul")).date(),
        "amount": abs(int(amount)),
        "category": category,
        "method": method,
        "memo": memo,
        "non_expense": non_expense,
    }
    quick_add_dialog()


@st.dialog("📝 빠른 입력 수정")
def quick_add_dialog():
    item = st.session_state.get("pending_quick_entry")    
    
    if not item:
        st.warning("등록할 항목이 없어요.")
        return

    current_cat = str(item["category"])
    cat_index = CATEGORY_OPTIONS.index(current_cat) if current_cat in CATEGORY_OPTIONS else 0

    current_method = str(item["method"]).strip() if str(item["method"]).strip() else DEFAULT_METHOD
    method_index = METHOD_OPTIONS.index(current_method) if current_method in METHOD_OPTIONS else 0

    base_memo, base_fuel_price, base_actual_amount, base_is_non_expense = split_fuel_memo(str(item["memo"]))

    with st.form("quick_add_edit_form"):
        q1, q2 = st.columns(2)

        with q1:
            category = st.selectbox(
                "카테고리",
                CATEGORY_OPTIONS,
                index=cat_index,
                key="quick_edit_cat"
            )

            d = st.date_input(
                "날짜",
                value=item["date"],
                key="quick_edit_date"
            )

            method = st.selectbox(
                "결제수단",
                METHOD_OPTIONS,
                index=method_index,
                key="quick_edit_method",
                disabled=False
            )

        with q2:
            memo = st.text_input(
                "메모",
                value=base_memo,
                key="quick_edit_memo"
            )

            amount_text = st.text_input(
                "금액",
                value=f"{abs(int(item['amount'])):,}",
                key="quick_edit_amount"
            )

            fuel_price = st.text_input(
                "리터당 가격",
                value=base_fuel_price,
                placeholder="주유일 때만 입력",
                key="quick_edit_fuel_price"
            )

        c_left, c_right = st.columns([1, 1])

        with c_right:
            non_expense = st.checkbox(
                "지출에 반영 안 함 (지출 제외)",
                value=False,
                key="quick_edit_non_expense"
            )

            if non_expense:
                st.caption("비지출 기록은 결제수단을 저장하지 않아요.")
                method = ""

        col_cancel, col_save = st.columns(2)
        with col_cancel:
            canceled = st.form_submit_button("취소", use_container_width=True)
        with col_save:
            saved = st.form_submit_button("💾 저장", use_container_width=True)

    if saved:
        amount_clean = amount_text.replace(",", "").strip()
        fuel_price_clean = fuel_price.replace(",", "").strip()

        if not amount_clean:
            st.error("금액을 입력해줘.")
        elif not re.fullmatch(r"\d+", amount_clean):
            st.error("금액은 숫자만 입력해줘.")
        elif fuel_price_clean and not re.fullmatch(r"\d+", fuel_price_clean):
            st.error("리터당 가격은 숫자만 입력해줘.")
        else:
            final_memo = build_fuel_memo(
                memo,
                fuel_price_clean,
                amount_clean,
                is_non_expense=non_expense
            )

            if non_expense:
                final_category = category
                final_method = ""
            else:
                final_category, final_method = auto_card_and_category(
                    final_memo,
                    category,
                    method or DEFAULT_METHOD
                )

            amount_value = int(amount_clean)

            if non_expense:
                final_amount = 0
            elif is_incident_income(final_method, final_memo):
                final_amount = amount_value
            else:
                final_amount = -amount_value

            row = {
                "date": str(d),
                "amount": final_amount,
                "category": final_category,
                "method": final_method,
                "memo": final_memo,
            }

            current_df = load_df()
            current_df = pd.concat([current_df, pd.DataFrame([row])], ignore_index=True)
            save_df(current_df)

            st.session_state.pending_quick_entry = None
            st.success("✅ 저장 완료!")
            st.rerun()

    if canceled:
        st.session_state.pending_quick_entry = None
        st.rerun()

def get_card_detail_df(month_df, method_name, detail_name):
    df = month_df[month_df["method"] == method_name].copy()

    if "date_dt" not in df.columns:
        df["date_dt"] = pd.to_datetime(df["date"], errors="coerce")
    
    # 통합 카테고리용
    if method_name == "통합" and detail_name == "미용":
        df = month_df.copy()

        if "date_dt" not in df.columns:
            df["date_dt"] = pd.to_datetime(df["date"], errors="coerce")

        # 현대카드 미용
        hyundai_beauty_df = df[
            (df["method"] == "현대카드") &
            (df["category"] == "미용") &
            (df["amount"] < 0)
        ].copy()

        # 사건비통장 미용
        incident_beauty_df = df[
            (df["method"] == "사건비통장") &
            (df["amount"] < 0)
        ].copy()

        if not incident_beauty_df.empty:
            incident_beauty_df["detail_category"] = incident_beauty_df["memo"].apply(classify_incident_memo)
            incident_beauty_df = incident_beauty_df[
                incident_beauty_df["detail_category"] == "미용"
            ].copy()

        merged_df = pd.concat([hyundai_beauty_df, incident_beauty_df], ignore_index=True)
        return merged_df.sort_values(by="date_dt", ascending=False)

    if method_name == "통합" and detail_name == "주유":
        df = month_df.copy()

        if "date_dt" not in df.columns:
            df["date_dt"] = pd.to_datetime(df["date"], errors="coerce")

        fuel_df = df[
            df["memo"].astype(str).str.contains("주유", na=False)
        ].copy()

        return fuel_df.sort_values(by="date_dt", ascending=False)

    # 사건비통장
    if method_name == "사건비통장":
        df = df[df["amount"] < 0].copy()

        if not df.empty:
            df["detail_category"] = df["memo"].apply(classify_incident_memo)
        else:
            df["detail_category"] = pd.Series(dtype="object")

        known = ["병원비", "약값", "검진", "선물", "경조사", "미용"]

        if detail_name == "기타":
            df = df[~df["detail_category"].isin(known)].copy()
        else:
            df = df[df["detail_category"] == detail_name].copy()

    # 현대카드
    elif method_name == "현대카드":
        known = ["쇼핑", "외식", "배달", "커피", "편의점"]

        df = df[df["amount"] < 0].copy()

        if detail_name == "기타":
            df = df[~df["category"].isin(known)].copy()
        else:
            df = df[df["category"] == detail_name].copy()

    # 신한카드
    elif method_name == "신한카드":
        df = df[df["amount"] < 0].copy()
        memo_series = df["memo"].astype(str)

        if detail_name == "주유":
            df = df[memo_series.str.contains("주유", na=False)].copy()
        elif detail_name == "통신비":
            df = df[memo_series.str.contains("통신비", na=False)].copy()
        elif detail_name == "인터넷":
            df = df[memo_series.str.contains("인터넷", na=False)].copy()
        elif detail_name == "쿠팡와우":
            df = df[memo_series.str.contains("쿠팡와우", na=False)].copy()
        elif detail_name == "이모티콘":
            df = df[memo_series.str.contains("이모티콘", na=False)].copy()
        elif detail_name == "기타":
            known_mask = (
                memo_series.str.contains("주유", na=False)
                | memo_series.str.contains("통신비", na=False)
                | memo_series.str.contains("인터넷", na=False)
                | memo_series.str.contains("쿠팡와우", na=False)
                | memo_series.str.contains("이모티콘", na=False)
            )
            df = df[~known_mask].copy()

    return df.sort_values(by="date_dt", ascending=False)

@st.dialog("📋 상세내역")
def card_detail_dialog():
    method_name = st.session_state.get("card_detail_method")
    detail_name = st.session_state.get("card_detail_name")

    if not method_name or not detail_name:
        st.info("선택된 항목이 없어요.")
        return

    detail_df = get_card_detail_df(month_df, method_name, detail_name)

    st.markdown(f"### {method_name} · {detail_name}")

    if detail_df.empty:
        st.write("내역이 없어요.")
        return

    # ⛽ 통합 주유 그래프/요약
    if method_name == "통합" and detail_name == "주유":
        fuel_stats_df = extract_fuel_stats_df(detail_df)

        if not fuel_stats_df.empty:
            total_fuel_amount = int(fuel_stats_df["fuel_amount"].fillna(0).sum())
            total_liters = float(fuel_stats_df["liters"].fillna(0).sum())

            valid_unit_price = fuel_stats_df["unit_price"].dropna()
            avg_unit_price = int(valid_unit_price.mean()) if not valid_unit_price.empty else 0

            mc1, mc2, mc3 = st.columns(3)
            with mc1:
                st.metric("총 주유금액", f"{total_fuel_amount:,}원")
            with mc2:
                st.metric("총 주유 리터", f"{total_liters:.2f}L")
            with mc3:
                st.metric("평균 리터당 가격", f"{avg_unit_price:,}원")

            chart_option = st.selectbox(
                "그래프 보기",
                ["리터당 가격", "주유금액", "주유리터"],
                key="fuel_chart_option_total"
            )

            chart_df = fuel_stats_df.copy()
            chart_df["날짜"] = chart_df["date_dt"].dt.strftime("%m-%d")

            if chart_option == "리터당 가격":
                st.line_chart(
                    chart_df.set_index("날짜")[["unit_price"]],
                    use_container_width=True
                )
            elif chart_option == "주유금액":
                st.line_chart(
                    chart_df.set_index("날짜")[["fuel_amount"]],
                    use_container_width=True
                )
            elif chart_option == "주유리터":
                st.line_chart(
                    chart_df.set_index("날짜")[["liters"]],
                    use_container_width=True
                )

            st.caption("비지출 주유까지 포함한 전체 주유 기록 기준이에요.")
        
    show_df = detail_df.copy()

    if method_name == "통합" and detail_name == "주유":
        fuel_stats_df = extract_fuel_stats_df(detail_df).reset_index(drop=True)
        show_df = show_df.reset_index(drop=True)
        show_df = pd.concat(
            [show_df, fuel_stats_df[["unit_price", "fuel_amount", "liters"]]],
            axis=1
        )

    show_df["날짜"] = show_df["date_dt"].dt.strftime("%Y-%m-%d")
    show_df["금액_num"] = show_df["amount"].abs().astype(int)
    show_df["금액"] = show_df["금액_num"].apply(lambda x: f"{x:,}원")

    cols = ["날짜"]

    if "method" in show_df.columns and method_name == "통합":
        cols.append("method")

    if "memo" in show_df.columns:
        cols.append("memo")

    if method_name == "통합" and detail_name == "주유":
        if "unit_price" in show_df.columns:
            show_df["unit_price_text"] = show_df["unit_price"].apply(
                lambda x: f"{int(x):,}원" if pd.notna(x) else "-"
            )
            cols.append("unit_price_text")

        if "liters" in show_df.columns:
            show_df["liters_text"] = show_df["liters"].apply(
                lambda x: f"{x:.2f}L" if pd.notna(x) else "-"
            )
            cols.append("liters_text")

    if method_name == "현대카드" and "category" in show_df.columns:
        cols.append("category")

    if method_name == "사건비통장" and "detail_category" in show_df.columns:
        cols.append("detail_category")

    if method_name == "통합":
        if "category" in show_df.columns:
            cols.append("category")
        if "detail_category" in show_df.columns:
            cols.append("detail_category")

    cols.append("금액")

    st.dataframe(
        show_df[cols].rename(columns={
            "method": "결제수단",
            "memo": "사용처/메모",
            "category": "카테고리",
            "detail_category": "세부분류",
            "unit_price_text": "리터당 가격",
            "liters_text": "주유량",            
        }),
        use_container_width=True,
        hide_index=True
    )

    st.markdown(
        f"<div style='text-align:right; font-weight:800; font-size:15px; margin:10px 0;'>"
        f"💰 합계: {int(show_df['금액_num'].sum()):,}원"
        f"</div>",
        unsafe_allow_html=True
    )

def render_card_detail_row(label, amount, method_name, key_suffix, display_label=None):
    if display_label is None:
        display_label = label

    col1, col2 = st.columns([3, 1])

    with col1:
        if st.button(
            display_label,
            key=f"card_detail_{method_name}_{key_suffix}",
            use_container_width=True,
            type="secondary"
        ):
            st.session_state["card_detail_method"] = method_name
            st.session_state["card_detail_name"] = label
            card_detail_dialog()

    with col2:
        st.markdown(
            f"<div style='text-align:right; padding-top:4px;'><b>{amount:,}원</b></div>",
            unsafe_allow_html=True
        )
        
# 공통 월 데이터
month_key = datetime.now(KST).strftime("%Y-%m")
month_df = df.copy()
month_df["date_dt"] = pd.to_datetime(month_df["date"], errors="coerce")
month_df = month_df[month_df["date_dt"].dt.strftime("%Y-%m") == month_key]

# 예산은 현대카드만
spent = int(
    month_df[month_df["method"] == BUDGET_METHOD]["amount"]
    .abs()
    .sum()
)
remaining = MONTHLY_BUDGET - spent
percent = min(spent / MONTHLY_BUDGET, 1.0) if MONTHLY_BUDGET > 0 else 0
status = budget_status(spent)

# 카드별 월 합계
card_raw_sum = month_df.groupby("method")["amount"].sum()

hyundai_amount = abs(int(card_raw_sum.get("현대카드", 0)))
shinhan_amount = abs(int(card_raw_sum.get("신한카드", 0)))

# 현대카드 세부내역
hyundai_df = month_df[month_df["method"] == "현대카드"].copy()

hyundai_shopping = abs(int(
    hyundai_df[hyundai_df["category"] == "쇼핑"]["amount"].sum()
))
hyundai_eatout = abs(int(
    hyundai_df[hyundai_df["category"] == "외식"]["amount"].sum()
))
hyundai_delivery = abs(int(
    hyundai_df[hyundai_df["category"] == "배달"]["amount"].sum()
))
hyundai_coffee = abs(int(
    hyundai_df[hyundai_df["category"] == "커피"]["amount"].sum()
))
hyundai_convenience = abs(int(
    hyundai_df[hyundai_df["category"] == "편의점"]["amount"].sum()
))
hyundai_beauty = abs(int(
    hyundai_df[hyundai_df["category"] == "미용"]["amount"].sum()
))

hyundai_known_total = (
    hyundai_shopping
    + hyundai_eatout
    + hyundai_delivery
    + hyundai_coffee
    + hyundai_convenience
    + hyundai_beauty
)

hyundai_other = max(hyundai_amount - hyundai_known_total, 0)

# -----------------------------
# 신한카드 세부내역
# -----------------------------
shinhan_df = month_df[month_df["method"] == "신한카드"].copy()

shinhan_fuel = abs(int(
    shinhan_df[shinhan_df["memo"].astype(str).str.contains("주유", na=False)]["amount"].sum()
))
shinhan_phone = abs(int(
    shinhan_df[shinhan_df["memo"].astype(str).str.contains("통신비", na=False)]["amount"].sum()
))
shinhan_internet = abs(int(
    shinhan_df[shinhan_df["memo"].astype(str).str.contains("인터넷", na=False)]["amount"].sum()
))
shinhan_wow = abs(int(
    shinhan_df[shinhan_df["memo"].astype(str).str.contains("쿠팡와우", na=False)]["amount"].sum()
))
shinhan_emoji = abs(int(
    shinhan_df[shinhan_df["memo"].astype(str).str.contains("이모티콘", na=False)]["amount"].sum()
))

shinhan_known_total = (
    shinhan_fuel
    + shinhan_phone
    + shinhan_internet
    + shinhan_wow
    + shinhan_emoji
)

shinhan_other = max(shinhan_amount - shinhan_known_total, 0)

fuel_all_df = month_df[
    month_df["memo"].astype(str).str.contains("주유", na=False)
].copy()

fuel_all_stats_df = extract_fuel_stats_df(fuel_all_df)

total_fuel_amount_all = int(fuel_all_stats_df["fuel_amount"].fillna(0).sum()) if not fuel_all_stats_df.empty else 0

# 사건비통장: 지출 / 환급 / 순금액
incident_df = month_df[month_df["method"] == "사건비통장"].copy()

incident_spent = abs(int(incident_df[incident_df["amount"] < 0]["amount"].sum()))
incident_refund = int(incident_df[incident_df["amount"] > 0]["amount"].sum())

# -----------------------------
# 사건비통장 세부 분류
# -----------------------------
incident_expense_df = incident_df[incident_df["amount"] < 0].copy()

if not incident_expense_df.empty:
    incident_expense_df["detail_category"] = incident_expense_df["memo"].apply(classify_incident_memo)
else:
    incident_expense_df["detail_category"] = pd.Series(dtype="object")

incident_hospital = abs(int(
    incident_expense_df[incident_expense_df["detail_category"] == "병원비"]["amount"].sum()
))

incident_medicine = abs(int(
    incident_expense_df[incident_expense_df["detail_category"] == "약값"]["amount"].sum()
))

incident_checkup = abs(int(
    incident_expense_df[incident_expense_df["detail_category"] == "검진"]["amount"].sum()
))

incident_gift = abs(int(
    incident_expense_df[incident_expense_df["detail_category"] == "선물"]["amount"].sum()
))

incident_event = abs(int(
    incident_expense_df[incident_expense_df["detail_category"] == "경조사"]["amount"].sum()
))

incident_beauty = abs(int(
    incident_expense_df[incident_expense_df["detail_category"] == "미용"]["amount"].sum()
))

incident_known_total = (
    incident_hospital
    + incident_medicine
    + incident_checkup
    + incident_gift
    + incident_event
    + incident_beauty
)

incident_other = max(incident_spent - incident_known_total, 0)

# 순금액
incident_amount = incident_spent - incident_refund

total_amount = hyundai_amount + shinhan_amount + incident_amount

total_beauty = hyundai_beauty + incident_beauty

with tab1:
    # -------------------
    # 상단 예산 현황
    # -------------------
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        render_budget_card(
            "이번달 예산",
            f"{MONTHLY_BUDGET:,}원",
            theme["container_bg"],
            theme["metric_border"],
            theme["button_text"]
        )
    
    with c2:
        render_budget_card(
            "지금까지 사용",
            f"{spent:,}원",
            theme["container_bg"],
            theme["metric_border"],
            theme["button_text"]
        )
    
    with c3:
        render_budget_card(
            "남은 금액",
            f"{remaining:,}원",
            theme["container_bg"],
            theme["metric_border"],
            theme["button_text"]
        )
    
    with c4:
        render_budget_card(
            "예산 상태",
            status["label"],
            status["bg"],
            status["border"],
            status["text"]
        )

    st.progress(percent)

    if spent > MONTHLY_BUDGET:
        st.warning(f"예산을 {spent - MONTHLY_BUDGET:,}원 초과했어요.")
    else:
        st.caption(f"{month_key} 예산 사용률: {percent * 100:.1f}%")

    # -------------------
    # 체크리스트
    # -------------------
    checklist_month_df = get_month_checklist(month_key)

    checked_count = int(checklist_month_df["checked"].sum())
    total_count = len(checklist_month_df)
    all_checked = total_count > 0 and checked_count == total_count

    checklist_month_df["amount"] = checklist_month_df["item"].apply(parse_checklist_amount)

    checked_amount = int(
        checklist_month_df.loc[checklist_month_df["checked"] == True, "amount"].sum()
    )
    total_amount = int(checklist_month_df["amount"].sum())

    expander_title = (
        f"✅ 이번달 체크리스트 ({checked_count}/{total_count}) "
        f"· {checked_amount:,} / {total_amount:,}원"
    )

    with st.expander(expander_title, expanded=not all_checked):
        check_cols = st.columns(2)

        checklist_df_reset = checklist_month_df.reset_index(drop=True)
        half = (len(checklist_df_reset) + 1) // 2

        left_items = checklist_df_reset.iloc[:half]
        right_items = checklist_df_reset.iloc[half:]

        with check_cols[0]:
            for _, row in left_items.iterrows():
                checked_now = st.checkbox(
                    row["item"],
                    value=bool(row["checked"]),
                    key=f"check_{month_key}_{row['item']}"
                )
                if checked_now != bool(row["checked"]):
                    update_checklist_item(month_key, row["item"], checked_now)
                    st.rerun()

        with check_cols[1]:
            for _, row in right_items.iterrows():
                checked_now = st.checkbox(
                    row["item"],
                    value=bool(row["checked"]),
                    key=f"check_{month_key}_{row['item']}"
                )
                if checked_now != bool(row["checked"]):
                    update_checklist_item(month_key, row["item"], checked_now)
                    st.rerun()

    st.divider()

    left_info, right_info = st.columns([1, 1.4])

    with left_info:
        st.subheader("🪙 이번달 무지출데이")

        no_spend_days = get_final_no_spend_days(df, month_key)
        no_spend_count = len(no_spend_days)

        from datetime import timedelta
        today = datetime.now(ZoneInfo("Asia/Seoul")).date()
        no_spend_set = set(no_spend_days)

        streak = 0
        for i in range(0, 30):
            check_day = today - timedelta(days=i)
            check_str = check_day.strftime("%Y-%m-%d")

            if check_str in no_spend_set:
                streak += 1
            else:
                break

        if streak == 0:
            msg = "🙂 오늘부터 다시 시작!"
        elif streak == 1:
            msg = "🔥 하루 성공중! 좀만 더!"
        elif streak == 2:
            msg = "🔥 이틀째 성공! 화이팅!"
        elif streak == 3:
            msg = "🔥 3일 연속 성공! 이얄~"
        else:
            msg = f"🔥 {streak}일 연속 성공중!!"

        with st.container(border=True):
            st.markdown(f"**🪙 무지출데이 {no_spend_count}일**")

            color = "#2F7A4A" if "성공" in msg else "#B4546A"

            st.markdown(
                f"""
                <div style='font-size:26px; font-weight:800; margin-bottom:calc(-1px + 1rem); color:{color};'>
                    {msg}
                </div>
                """,
                unsafe_allow_html=True
            )

        if no_spend_days:
            recent_no_spend = sorted(no_spend_days)[-7:]
            recent_no_spend_short = [
                f"{int(d.split('-')[1])}/{int(d.split('-')[2])}" for d in recent_no_spend
            ]
            st.caption("📅 최근: " + ", ".join(recent_no_spend_short))
        else:
            st.caption("아직 기록된 무지출데이가 없어요.")

    with right_info:
        st.subheader("📅 월별 예산 결산")

        review_df = get_monthly_budget_reviews(df, months=6)

        if review_df.empty:
            st.caption("아직 결산할 데이터가 없어요.")
        else:
            for _, r in review_df.iterrows():
                st.markdown(
                    f"""
                    <div style="
                        background:{r['bg']};
                        border:1px solid {r['border']};
                        border-radius:18px;
                        padding:14px 16px;
                        margin-bottom:10px;
                    ">
                        <div style="font-size:14px; font-weight:800; color:{r['text']}; margin-bottom:6px;">
                            {r['month']} 결산
                        </div>
                        <div style="font-size:18px; font-weight:800; color:{r['text']}; margin-bottom:4px;">
                            사용 {int(r['spent']):,}원 · {r['result_text']}
                        </div>
                        <div style="font-size:14px; color:{r['text']};">
                            {r['comment']}
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

    st.divider()

    # -------------------
    # 카드별 이번달 사용
    # -------------------
    st.subheader("💳 카드별 이번달 사용")

    card_col1, card_col2, card_col3, card_col4 = st.columns(4)

    with card_col1:
        render_budget_card(
            "💳 현대카드(용돈)",
            f"{hyundai_amount:,}원",
            "#F4F4F4",
            "#D6D6D6",
            "#4A4A4A"
        )

        st.markdown("<div style='margin-top:10px;'></div>", unsafe_allow_html=True)

        with st.container(border=True):

            if hyundai_shopping > 0:
                render_card_detail_row("쇼핑", hyundai_shopping, "현대카드", "shopping", "🛒 쇼핑")
            if hyundai_eatout > 0:
                render_card_detail_row("외식", hyundai_eatout, "현대카드", "eatout", "🍚 외식")
            if hyundai_delivery > 0:
                render_card_detail_row("배달", hyundai_delivery, "현대카드", "delivery", "🛵 배달")
            if hyundai_coffee > 0:
                render_card_detail_row("커피", hyundai_coffee, "현대카드", "coffee", "☕ 커피")
            if hyundai_convenience > 0:
                render_card_detail_row("편의점", hyundai_convenience, "현대카드", "convenience", "🏪 편의점")
            if hyundai_beauty > 0:
                render_card_detail_row("미용", hyundai_beauty, "현대카드", "beauty", "💅 미용")
            if hyundai_other > 0:
                render_card_detail_row("기타", hyundai_other, "현대카드", "other", "🧾 기타")

    with card_col2:
        render_budget_card(
            "🏦 신한카드 (고정비)",
            f"{shinhan_amount:,}원",
            "#F2FBFF",
            "#BFE8F7",
            "#3E7C91"
        )

        st.markdown("<div style='margin-top:10px;'></div>", unsafe_allow_html=True)

        with st.container(border=True):

            if shinhan_fuel > 0:
                render_card_detail_row("주유", shinhan_fuel, "신한카드", "fuel", "⛽ 주유")
            if shinhan_phone > 0:
                render_card_detail_row("통신비", shinhan_phone, "신한카드", "phone", "📱 통신비")
            if shinhan_internet > 0:
                render_card_detail_row("인터넷", shinhan_internet, "신한카드", "internet", "🌐 인터넷")
            if shinhan_wow > 0:
                render_card_detail_row("쿠팡와우", shinhan_wow, "신한카드", "wow", "📦 쿠팡와우")
            if shinhan_emoji > 0:
                render_card_detail_row("이모티콘", shinhan_emoji, "신한카드", "emoji", "💬 이모티콘")
            if shinhan_other > 0:
                render_card_detail_row("기타", shinhan_other, "신한카드", "other", "🧾 기타")

    with card_col3:
        render_budget_card(
            "🧾 사건비통장 (특수지출)",
            f"{incident_amount:,}원",
            "#FFF8CC",
            "#F2E18B",
            "#8A6A00"
        )

        st.markdown("<div style='margin-top:10px;'></div>", unsafe_allow_html=True)

        with st.container(border=True):
            if incident_hospital > 0:
                render_card_detail_row("병원비", incident_hospital, "사건비통장", "hospital", "🏥 병원비")
            if incident_medicine > 0:
                render_card_detail_row("약값", incident_medicine, "사건비통장", "medicine", "💊 약값")
            if incident_checkup > 0:
                render_card_detail_row("검진", incident_checkup, "사건비통장", "checkup", "🩺 검진")
            if incident_beauty > 0:
                render_card_detail_row("미용", incident_beauty, "사건비통장", "beauty", "💅 미용")
            if incident_gift > 0:
                render_card_detail_row("선물", incident_gift, "사건비통장", "gift", "🎁 선물")
            if incident_event > 0:
                render_card_detail_row("경조사", incident_event, "사건비통장", "event", "🙏 경조사")            
            if incident_other > 0:
                render_card_detail_row("기타", incident_other, "사건비통장", "other", "🧾 기타")
            
            st.markdown(
                f"<div style='text-align:right; font-size:13px; opacity:0.7; margin-bottom:10px;'>"
                f"지출 {incident_spent:,}원 / 💰 환급 {incident_refund:,}원"
                f"</div>",
                unsafe_allow_html=True
            )

    with card_col4:
        render_budget_card(
            "이번달 총 지출",
            f"{total_amount:,}원",
            "#FFEFF6",
            "#FFC4D6",
            "#A85E74"
        )

        st.markdown("<div style='margin-top:10px;'></div>", unsafe_allow_html=True)
        
        with st.container(border=True):
            render_card_detail_row("미용", total_beauty, "통합", "total_beauty", "💅 미용 총 지출")

            if total_fuel_amount_all > 0:
                render_card_detail_row("주유", total_fuel_amount_all, "통합", "total_fuel", "⛽ 주유 총 지출")

    st.divider()

    # =========================
    # 날짜 선택 입력
    # =========================
    st.subheader("🗓 날짜 선택해서 입력")

    with st.form("manual_add", clear_on_submit=True):
        m1, m2, m3 = st.columns(3)

        with m1:
            d = st.date_input("날짜", value=datetime.now(ZoneInfo("Asia/Seoul")).date(), key="manual_date")
            category = st.selectbox(
                "카테고리",
                CATEGORY_OPTIONS,
                index=CATEGORY_OPTIONS.index("쇼핑"),
                key="manual_cat"
            )
            no_spend_only = st.checkbox(
                "이 날짜를 무지출데이로 기록",
                value=False,
                key="manual_no_spend_only"
            )
            
        with m2:
            memo = st.text_input("메모", value="", key="manual_memo")
            amount_text = st.text_input("금액", value="", placeholder="금액 입력", key="manual_amount")

        with m3:
            fuel_price = st.text_input(
                "리터당 가격",
                value="",
                placeholder="주유일 때만 입력",
                key="manual_fuel_price"
            )

            method = st.selectbox(
                "결제수단",
                METHOD_OPTIONS,
                index=METHOD_OPTIONS.index(DEFAULT_METHOD),
                key="manual_method"
            )

            non_expense = st.checkbox(
                "지출에 반영 안 함 (비지출 기록)",
                value=False,
                key="manual_non_expense"
            )

            if non_expense:
                st.caption("비지출 기록은 결제수단을 저장하지 않아요.")
                method = ""

        submitted_manual = st.form_submit_button("추가", use_container_width=True)

    if submitted_manual:
        if no_spend_only:
            update_no_spend_day(str(d), True)
            st.success(f"✅ {str(d)} 무지출데이로 기록했어요!")
            st.rerun()

        amount_clean = amount_text.replace(",", "").strip()
        fuel_price_clean = fuel_price.replace(",", "").strip()

        if not amount_clean:
            st.error("금액을 입력해줘.")
        elif not re.fullmatch(r"\d+", amount_clean):
            st.error("금액은 숫자만 입력해줘. 예: 4500 또는 4,500")
        elif fuel_price_clean and not re.fullmatch(r"\d+", fuel_price_clean):
            st.error("리터당 가격은 숫자만 입력해줘.")
        else:
            final_memo = build_fuel_memo(
                memo,
                fuel_price_clean,
                amount_clean,
                is_non_expense=non_expense
            )

            if non_expense:
                final_category = category
                final_method = ""
            else:
                final_category, final_method = auto_card_and_category(
                    final_memo,
                    category,
                    method or DEFAULT_METHOD
                )

            amount_value = int(amount_clean)

            if non_expense:
                final_amount = 0
            elif is_incident_income(final_method, final_memo):
                final_amount = amount_value
            else:
                final_amount = -amount_value

            row = {
                "date": str(d),
                "amount": final_amount,
                "category": final_category,
                "method": final_method,
                "memo": final_memo,
            }
            df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
            save_df(df)
            st.success("✅ 저장 완료!")
            st.rerun()
    
    st.divider()

    # =========================
    # 입력예시 | 빠른입력
    # =========================
    left_col, right_col = st.columns([1, 1])

    with left_col:
        st.subheader("📝 입력 예시 / 규칙")
        st.markdown("""
- 빠른 입력: `금액 메모 @결제수단`
- 예: `12000 점심 @현대카드`
- `우유 1000` / `1000 우유` 둘 다 가능
- 날짜 지정: `YYYY-MM-DD`를 맨 앞에
- 예: `2026-03-01 4500 스타벅스 @현금`
- 사건비통장에서 `환급`, `입금`, `수입`, `보험금` 단어가 들어가면 자동 수입 처리
- 지정 키워드가 없으면 카테고리는 기본 `쇼핑`
""")

    with right_col:
        st.subheader("⚡ 빠른 입력")
        with st.form("quick_add", clear_on_submit=True):
            quick_category = st.selectbox(
                "카테고리",
                CATEGORY_OPTIONS,
                index=CATEGORY_OPTIONS.index("쇼핑"),
                key="quick_category_select"
            )
        
            quick = st.text_input(
                "입력",
                placeholder="4500 스타벅스 @현대카드",
                key="quick_input_text"
            )
        
            submitted_quick = st.form_submit_button("저장 (Enter)", use_container_width=True)        

        if submitted_quick:
            try:
                row = parse_quick_input(
                    quick,
                    default_category=quick_category,
                    default_method=DEFAULT_METHOD
                )
                df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
                save_df(df)
                st.success("✅ 저장 완료!")
                st.rerun()
            except Exception as e:
                st.error(f"저장 실패: {e}")

    st.divider()

    # =========================
    # 빠른 입력 보조
    # =========================
    with st.expander("⚡ 빠른 입력 보조", expanded=False):
        top_left, top_right = st.columns([1, 1])

        with top_left:
            st.subheader("⚡ 자주 쓰는 버튼")

            b1, b2 = st.columns(2)
            b3, b4 = st.columns(2)

            with b1:
                if st.button("☕ 커피 4,500", use_container_width=True, key="btn_coffee"):
                    open_quick_edit(4500, "커피")

            with b2:
                if st.button("🍚 외식 12,000", use_container_width=True, key="btn_eatout"):
                    open_quick_edit(12000, "외식")

            with b3:
                if st.button("🛵 배달 18,000", use_container_width=True, key="btn_delivery"):
                    open_quick_edit(18000, "배달")

            with b4:
                if st.button("🛒 쇼핑 30,000", use_container_width=True, key="btn_shopping"):
                    open_quick_edit(30000, "쇼핑")

        with top_right:
            st.subheader("📌 신한카드 고정비")

            shinhan_fixed_buttons = [
                ("📱 통신비", 103490, "통신비"),
                ("🌐 인터넷", 26400, "인터넷"),
                ("📦 쿠팡와우", 7890, "쿠팡와우"),
                ("💬 이모티콘", 3900, "이모티콘"),
                ("⛽ 주유", 65000, "주유"),
            ]

            cols_per_row = 2

            for i in range(0, len(shinhan_fixed_buttons), cols_per_row):
                cols = st.columns(cols_per_row)

                for j, (label, amount, memo) in enumerate(shinhan_fixed_buttons[i:i+cols_per_row]):
                    with cols[j]:
                        if st.button(label, use_container_width=True, key=f"btn_{memo}"):
                            open_quick_edit(amount, "고정비", memo, "신한카드")
    
    with st.expander("11111", expanded=False):
        st.subheader("22222")
        st.write("33333")
 
    st.divider()

with tab2:
    # =========================
    # 검색 / 월선택 / 다운로드
    # =========================
    top_filter_left, top_filter_right = st.columns([1, 1])

    with top_filter_left:
        KST = ZoneInfo("Asia/Seoul")

        month_options = get_month_options(df)
        current_month = datetime.now(KST).strftime("%Y-%m")

        if "selected_month" not in st.session_state or st.session_state["selected_month"] not in month_options:
            st.session_state["selected_month"] = current_month

        month = st.selectbox("월 선택", month_options, key="selected_month")
        q = st.text_input(
            "검색(카테고리/메모/결제수단)",
            placeholder="예: 스타벅스 / 배달 / 현대카드 / 환급",
            key="search_text"
        )

        with top_filter_right:
            download_month = st.selectbox(
                "다운로드할 월 선택",
                month_options,
                key="download_month"
            )    
        
            # 다운로드용 데이터
            download_view = df.copy()
            download_view["date_dt"] = pd.to_datetime(download_view["date"], errors="coerce")
        
            try:
                dy, dm = download_month.split("-")
                download_view = download_view[
                    (download_view["date_dt"].dt.year == int(dy)) &
                    (download_view["date_dt"].dt.month == int(dm))
                ]
            except Exception:
                download_view = download_view.iloc[0:0]
        
            if download_view.empty:
                st.caption("선택한 월의 다운로드할 내역이 없어요.")
            else:
                download_df = download_view.drop(columns=["date_dt"], errors="ignore").copy()
                download_df = download_df.sort_values(by="date", ascending=False)

                # 구분 추가
                download_df["구분"] = download_df["amount"].apply(
                    lambda x: "환급" if int(x) > 0 else "지출"
                )

                # 숫자 보관용
                download_df["amount_num"] = download_df["amount"].abs()

                # 표시용 금액
                download_df["금액"] = download_df["amount_num"].apply(lambda x: f"{x:,}원")

                # 한글 컬럼으로 정리
                download_df["날짜"] = download_df["date"]
                download_df["카테고리"] = download_df["category"]
                download_df["결제수단"] = download_df["method"]
                download_df["메모"] = download_df["memo"]

                # 번호 생성
                download_df = download_df.reset_index(drop=True)
                download_df.insert(0, "번호", range(1, len(download_df) + 1))

                # 최종 컬럼 순서
                download_df = download_df[
                    ["번호", "날짜", "구분", "카테고리", "메모", "금액", "결제수단", "amount_num"]
                ]

                buffer = BytesIO()

                with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    download_df.drop(columns=["amount_num"]).to_excel(
                        writer,
                        index=False,
                        sheet_name="가계부"
                    )

                    worksheet = writer.sheets["가계부"]

                    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

                    # 스타일 정의
                    header_font = Font(bold=True, color="FFFFFF")
                    bold_font = Font(bold=True)

                    header_fill = PatternFill("solid", fgColor="B96A5C")
                    summary_fill = PatternFill("solid", fgColor="FDECEC")
                    refund_fill = PatternFill("solid", fgColor="EAF4FF")
                    expense_fill = PatternFill("solid", fgColor="FFF4F4")

                    thin_border = Border(
                        left=Side(style="thin", color="E5D6D1"),
                        right=Side(style="thin", color="E5D6D1"),
                        top=Side(style="thin", color="E5D6D1"),
                        bottom=Side(style="thin", color="E5D6D1"),
                    )

                    center_align = Alignment(horizontal="center", vertical="center")
                    right_align = Alignment(horizontal="right", vertical="center")
                    left_align = Alignment(horizontal="left", vertical="center")

                    # 헤더 스타일
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_align
                        cell.border = thin_border

                    # 열 너비
                    column_widths = [8, 12, 10, 12, 28, 12, 12]  # 번호, 날짜, 구분, 카테고리, 메모, 금액, 결제수단
                    for i, width in enumerate(column_widths, start=1):
                        worksheet.column_dimensions[chr(64 + i)].width = width

                    # 본문 스타일
                    data_row_count = len(download_df)
                    for row_idx in range(2, data_row_count + 2):
                        row_type = worksheet[f"C{row_idx}"].value  # 구분 컬럼

                        for col_idx in range(1, 8):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.border = thin_border

                            if col_idx in [1, 2, 3, 4, 7]:
                                cell.alignment = center_align
                            elif col_idx == 6:
                                cell.alignment = right_align
                            else:
                                cell.alignment = left_align

                        # 지출 / 환급 행 색상
                        if row_type == "환급":
                            for col_idx in range(1, 8):
                                worksheet.cell(row=row_idx, column=col_idx).fill = refund_fill
                        else:
                            for col_idx in range(1, 8):
                                worksheet.cell(row=row_idx, column=col_idx).fill = expense_fill

                    # 합계 계산
                    spent_total = download_df.loc[download_df["구분"] == "지출", "amount_num"].sum()
                    refund_total = download_df.loc[download_df["구분"] == "환급", "amount_num"].sum()
                    net_total = spent_total - refund_total

                    start_row = len(download_df) + 3

                    summary_rows = [
                        ("지출 합계", f"{spent_total:,}원"),
                        ("환급 합계", f"{refund_total:,}원"),
                        ("순지출", f"{net_total:,}원"),
                    ]

                    for i, (label, value) in enumerate(summary_rows):
                        row_no = start_row + i
                        worksheet[f"E{row_no}"] = label
                        worksheet[f"F{row_no}"] = value

                        worksheet[f"E{row_no}"].font = bold_font
                        worksheet[f"F{row_no}"].font = bold_font

                        worksheet[f"E{row_no}"].fill = summary_fill
                        worksheet[f"F{row_no}"].fill = summary_fill

                        worksheet[f"E{row_no}"].alignment = center_align
                        worksheet[f"F{row_no}"].alignment = right_align

                        worksheet[f"E{row_no}"].border = thin_border
                        worksheet[f"F{row_no}"].border = thin_border

                    # 카테고리 합계 시트
                    category_sum = (
                        download_df
                        .groupby("카테고리")["amount_num"]
                        .sum()
                        .reset_index()
                    )

                    category_sum["합계"] = category_sum["amount_num"].apply(lambda x: f"{x:,}원")
                    category_sum = category_sum[["카테고리", "합계"]]

                    category_sum.to_excel(
                        writer,
                        index=False,
                        sheet_name="카테고리합계"
                    )

                    sheet2 = writer.sheets["카테고리합계"]

                    # 카테고리합계 헤더 스타일
                    for cell in sheet2[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_align
                        cell.border = thin_border

                    sheet2.column_dimensions["A"].width = 15
                    sheet2.column_dimensions["B"].width = 15

                    for row in sheet2.iter_rows(min_row=2, max_col=2):
                        row[0].alignment = center_align
                        row[1].alignment = right_align
                        row[0].border = thin_border
                        row[1].border = thin_border

                excel_data = buffer.getvalue()

                st.download_button(
                    label=f"📥 {download_month} 가계부 다운로드",
                    data=excel_data,
                    file_name=f"{download_month}_가계부.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

    st.divider()

    if "record_view_mode" not in st.session_state:
        st.session_state["record_view_mode"] = "calendar"

    view_mode = st.radio(
        "보기 방식",
        options=["list", "calendar"],
        format_func=lambda x: "📋 리스트" if x == "list" else "🗓 달력",
        horizontal=True,
        key="record_view_mode"
    )

    calendar_df = df.copy()
    calendar_df["date_dt"] = pd.to_datetime(calendar_df["date"], errors="coerce")

    if view_mode == "calendar":
        st.subheader("🗓 월별 달력 보기")

        render_month_calendar(calendar_df, month, theme)

        st.markdown("<div style='margin-top:20px;'></div>", unsafe_allow_html=True)

        calendar_summary_df = get_calendar_day_summary(calendar_df, month)
        calendar_no_spend_days = get_final_no_spend_days(calendar_df, month)

        cal_spent = int(calendar_summary_df["spent"].sum()) if not calendar_summary_df.empty else 0
        cal_refund = int(calendar_summary_df["refund"].sum()) if not calendar_summary_df.empty else 0
        cal_no_spend = len(calendar_no_spend_days)
        cal_net = cal_spent - cal_refund

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.metric("월 지출 합계", f"{cal_spent:,}원")
        with c2:
            st.metric("월 환급 합계", f"{cal_refund:,}원")
        with c3:
            st.metric("월 순지출", f"{cal_net:,}원")
        with c4:
            st.metric("무지출데이", f"{cal_no_spend}일")

        st.caption("💡 초록 배경은 무지출데이, ⛽는 주유 기록이 있는 날이에요.")

        st.divider()
        st.subheader("📅 날짜별 세부내역")

        KST = ZoneInfo("Asia/Seoul")
        today_date = datetime.now(KST).date()
        today_month_str = today_date.strftime("%Y-%m")

        default_calendar_date = (
            today_date if month == today_month_str else pd.to_datetime(f"{month}-01").date()
        )

        if "calendar_detail_date" not in st.session_state:
            st.session_state["calendar_detail_date"] = default_calendar_date
        else:
            current_selected = st.session_state["calendar_detail_date"]
            if current_selected.strftime("%Y-%m") != month:
                st.session_state["calendar_detail_date"] = default_calendar_date

        selected_calendar_date = st.date_input(
            "날짜 선택",
            key="calendar_detail_date"
        )

        selected_date_str = selected_calendar_date.strftime("%Y-%m-%d")

        day_detail_df = calendar_df.copy()
        day_detail_df = day_detail_df[
            day_detail_df["date_dt"].dt.strftime("%Y-%m") == month
        ].copy()
        day_detail_df = day_detail_df[
            day_detail_df["date_dt"].dt.date == selected_calendar_date
        ].copy()

        manual_no_spend_df = load_no_spend_df()
        manual_no_spend_checked = False
        if not manual_no_spend_df.empty:
            manual_no_spend_df["date"] = manual_no_spend_df["date"].astype(str)
            manual_no_spend_checked = bool(
                ((manual_no_spend_df["date"] == selected_date_str) &
                 (manual_no_spend_df["checked"] == True)).any()
            )

        auto_no_spend_days = get_auto_no_spend_days(calendar_df, month)
        is_auto_no_spend = selected_date_str in auto_no_spend_days
        is_final_no_spend = selected_date_str in calendar_no_spend_days

        st.caption(f"선택 날짜: {selected_date_str}")

        if day_detail_df.empty:
            if is_final_no_spend:
                if manual_no_spend_checked and not is_auto_no_spend:
                    st.success(f"{selected_date_str} · 🪙 수동으로 기록한 무지출데이예요!")
                else:
                    st.success(f"{selected_date_str} · 🪙 자동 집계된 무지출데이예요!")
            else:
                st.info(f"{selected_date_str} 내역이 없어요.")
        else:
            show_day_df = day_detail_df.copy()
            show_day_df["날짜"] = show_day_df["date_dt"].dt.strftime("%Y-%m-%d")
            show_day_df["금액_num"] = show_day_df["amount"].abs().astype(int)
            show_day_df["금액"] = show_day_df["금액_num"].apply(lambda x: f"{x:,}원")

            day_cols = ["날짜", "category", "memo", "method", "금액"]

            st.dataframe(
                show_day_df[day_cols].rename(columns={
                    "category": "카테고리",
                    "memo": "메모",
                    "method": "결제수단",
                }),
                use_container_width=True,
                hide_index=True
            )

            day_spent = int(show_day_df[show_day_df["amount"] < 0]["amount"].abs().sum())
            day_refund = int(show_day_df[show_day_df["amount"] > 0]["amount"].sum())
            day_net = day_spent - day_refund

            st.markdown(
                f"""
                <div style="text-align:right; font-size:14px; font-weight:700; margin-top:8px;">
                    💸 지출 {day_spent:,}원 &nbsp;&nbsp; | &nbsp;&nbsp;
                    💰 환급 {day_refund:,}원 &nbsp;&nbsp; | &nbsp;&nbsp;
                    📊 순지출 {day_net:,}원
                </div>
                """,
                unsafe_allow_html=True
            )

            if is_final_no_spend:
                if manual_no_spend_checked and not is_auto_no_spend:
                    st.caption("🪙 이 날짜는 수동으로 무지출데이 추가한 날이에요.")
                else:
                    st.caption("🪙 이 날짜는 자동 집계된 무지출데이예요.")

        st.divider()

    current_view_key = f"{month}|{q}|{st.session_state.get('record_filter', '전체')}"
    if "last_view_key" not in st.session_state:
        st.session_state["last_view_key"] = current_view_key

    if st.session_state["last_view_key"] != current_view_key:
        st.session_state["record_page"] = 1
        st.session_state["last_view_key"] = current_view_key

    if view_mode == "list":
        # =========================
        # 기록 보기 필터
        # =========================
        st.subheader("🧾 기록 보기 / 필터")
        
        if "record_filter" not in st.session_state:
            st.session_state["record_filter"] = "전체"
        
        filter_items = [
            ("전체", "전체"),
            ("💳현대", "현대"),
            ("💳신한", "신한"),
            ("⛽주유", "주유"),
            ("📂사건비", "사건비"),
            ("💰환급", "환급"),
        ]
        
        cols = st.columns(len(filter_items))
        
        for i, (label, value) in enumerate(filter_items):
            with cols[i]:
                if st.button(
                    label,
                    use_container_width=True,
                    key=f"filter_{value}",
                    type="primary" if st.session_state["record_filter"] == value else "secondary"
                ):
                    st.session_state["record_filter"] = value
                    st.session_state["record_page"] = 1
                    st.rerun()
        
        record_filter = st.session_state["record_filter"]
        
        # 보기용 데이터
        view = df.copy()
        view["date_dt"] = pd.to_datetime(view["date"], errors="coerce")
        
        try:
            y, m = month.split("-")
            view = view[
                (view["date_dt"].dt.year == int(y)) &
                (view["date_dt"].dt.month == int(m))
            ]
        except Exception:
            pass
        
        if q.strip():
            qq = q.strip().lower()
            mask = (
                view["category"].astype(str).str.lower().str.contains(qq, na=False)
                | view["memo"].astype(str).str.lower().str.contains(qq, na=False)
                | view["method"].astype(str).str.lower().str.contains(qq, na=False)
            )
            view = view[mask]
        
        view = view.sort_values(by=["date_dt"], ascending=False)
        
        if record_filter == "현대":
            view = view[view["method"] == "현대카드"]
        
        elif record_filter == "신한":
            view = view[view["method"] == "신한카드"]
        
        elif record_filter == "주유":
            view = view[view["memo"].astype(str).str.contains("주유", na=False)]
        
        elif record_filter == "사건비":
            view = view[view["category"] == "사건비"]
        
        elif record_filter == "환급":
            view = view[(view["method"] == "사건비통장") & (view["amount"] > 0)]
        
        spent_total = int(view[view["amount"] < 0]["amount"].abs().sum())
        refund_total = int(view[view["amount"] > 0]["amount"].sum())
        net_total = spent_total - refund_total

        st.markdown(
            f"<div style='text-align:right; font-size:13px; opacity:0.75;'>"
            f"🔎 현재 보기: {record_filter} &nbsp;&nbsp; | &nbsp;&nbsp; "
            f"💸 지출 {spent_total:,}원 &nbsp;&nbsp; | &nbsp;&nbsp; "
            f"💰 환급 {refund_total:,}원 &nbsp;&nbsp; | &nbsp;&nbsp; "
            f"📊 순지출 {net_total:,}원"
            f"</div>",
            unsafe_allow_html=True
        )
        
        if view.empty:
            st.write("표시할 기록이 없어요.")
        else:
            view_with_idx = view.copy()
            view_with_idx["row_id"] = view_with_idx.index
            view_with_idx = view_with_idx.reset_index(drop=True)
        
            total_rows = len(view_with_idx)
            total_pages = (total_rows - 1) // PAGE_SIZE + 1
        
            if "record_page" not in st.session_state:
                st.session_state["record_page"] = 1
        
            # 페이지 범위 보정
            if st.session_state["record_page"] > total_pages:
                st.session_state["record_page"] = total_pages
            if st.session_state["record_page"] < 1:
                st.session_state["record_page"] = 1
        
            start_idx = (st.session_state["record_page"] - 1) * PAGE_SIZE
            end_idx = start_idx + PAGE_SIZE
        
            page_view = view_with_idx.iloc[start_idx:end_idx].copy()
            page_view["no"] = range(start_idx + 1, min(end_idx, total_rows) + 1)

            st.markdown("<br>", unsafe_allow_html=True)
            
            st.markdown(
                f"<p style='text-align:right; font-size:13px; opacity:0.75;'>"
                f"총 {total_rows}건"
                f"</p>",
                unsafe_allow_html=True
            )
        
            h0, h1, h2, h3, h4, h5, h6, h7 = st.columns([0.6, 1.1, 1.2, 2.2, 1.2, 1.1, 0.8, 0.8])
            h0.markdown("<div class='table-head'>번호</div>", unsafe_allow_html=True)
            h1.markdown("<div class='table-head'>날짜</div>", unsafe_allow_html=True)
            h2.markdown("<div class='table-head'>카테고리</div>", unsafe_allow_html=True)
            h3.markdown("<div class='table-head'>메모</div>", unsafe_allow_html=True)
            h4.markdown("<div class='table-head'>금액</div>", unsafe_allow_html=True)
            h5.markdown("<div class='table-head'>결제수단</div>", unsafe_allow_html=True)
            h6.markdown("<div class='table-head'>삭제</div>", unsafe_allow_html=True)
            h7.markdown("<div class='table-head'>수정</div>", unsafe_allow_html=True)
        
            for _, r in page_view.iterrows():
                c0, c1, c2, c3, c4, c5, c6, c7 = st.columns([0.6, 1.1, 1.2, 2.2, 1.2, 1.1, 0.8, 0.8])
        
                c0.markdown(f"<div class='row-box'>{r['no']}</div>", unsafe_allow_html=True)
                c1.markdown(f"<div class='row-box'>{r['date']}</div>", unsafe_allow_html=True)
                c2.markdown(f"<div class='row-box'><span class='cat-tag'>{r['category']}</span></div>", unsafe_allow_html=True)
                c3.markdown(f"<div class='row-box'>{r['memo']}</div>", unsafe_allow_html=True)
        
                memo_text = str(r["memo"])
                actual_match = re.search(r"실제\s*([\d,]+)원", memo_text)

                if "[비지출]" in memo_text and actual_match:
                    actual_display = f"{actual_match.group(1)}원"
                    amount_html = f"<div class='row-box amount-text'>⛽ 비지출 · {actual_display}</div>"
                else:
                    amount_display = f"{abs(int(r['amount'])):,}원"
                    if int(r["amount"]) > 0:
                        amount_html = f"<div class='row-box amount-text'>➕ {amount_display}</div>"
                    else:
                        amount_html = f"<div class='row-box amount-text'>💸 {amount_display}</div>"
        
                c4.markdown(amount_html, unsafe_allow_html=True)
        
                method = str(r["method"]).strip()

                if not method:
                    method_html = "<span style='opacity:0.5;'>-</span>"
                elif method == "신한카드":
                    method_html = f"<span class='method-shinhan'>{method}</span>"
                elif method == "사건비통장":
                    method_html = f"<span class='method-incident'>{method}</span>"
                else:
                    method_html = f"<span class='method-hyundai'>{method}</span>"
        
                c5.markdown(f"<div class='row-box'>{method_html}</div>", unsafe_allow_html=True)
        
                rid = int(r["row_id"])
        
                with c6:
                    if st.button("🗑", key=f"del_{rid}", use_container_width=True):
                        df = df.drop(index=rid).reset_index(drop=True)
                        save_df(df)
                        st.success("삭제 완료!")
                        st.rerun()
        
                with c7:
                    if st.button("✏", key=f"edit_{rid}", use_container_width=True):
                        edit_dialog(rid)
        
            st.markdown("<br>", unsafe_allow_html=True)
            
            p1, p2, p3 = st.columns([1.2, 5, 1.2])
            
            current_page = st.session_state["record_page"]
            
            # 표시할 페이지 번호 계산
            pages_to_show = []
            
            if total_pages <= 10:
                pages_to_show = list(range(1, total_pages + 1))
            else:
                if current_page <= 4:
                    pages_to_show = [1, 2, 3, 4, 5, "...", total_pages]
                elif current_page >= total_pages - 3:
                    pages_to_show = [1, "...", total_pages - 4, total_pages - 3, total_pages - 2, total_pages - 1, total_pages]
                else:
                    pages_to_show = [1, "...", current_page - 1, current_page, current_page + 1, "...", total_pages]
            
            # 이전 버튼
            with p1:
                prev_disabled = current_page <= 1
                if st.button(
                    "◀ 이전",
                    use_container_width=True,
                    key="prev_page",
                    disabled=prev_disabled
                ):
                    st.session_state["record_page"] -= 1
                    st.rerun()
            
            # 숫자 페이지
            with p2:
                page_cols = st.columns(len(pages_to_show))
            
                for i, page_item in enumerate(pages_to_show):
                    with page_cols[i]:
                        if page_item == "...":
                            st.markdown(
                                "<div style='text-align:center; padding-top:8px; font-weight:700;'>...</div>",
                                unsafe_allow_html=True
                            )
                        else:
                            if st.button(
                                str(page_item),
                                use_container_width=True,
                                key=f"page_{page_item}",
                                type="primary" if current_page == page_item else "secondary"
                            ):
                                st.session_state["record_page"] = page_item
                                st.rerun()
            
            # 다음 버튼
            with p3:
                next_disabled = current_page >= total_pages
                if st.button(
                    "다음 ▶",
                    use_container_width=True,
                    key="next_page",
                    disabled=next_disabled
                ):
                    st.session_state["record_page"] += 1
                    st.rerun()
                    
        st.divider()

with tab3:
    render_living_tab(get_worksheet, render_budget_card)
